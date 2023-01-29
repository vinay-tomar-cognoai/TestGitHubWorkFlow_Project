from CampaignApp.models import VoiceCampaignHistoryExportRequest, Campaign, CampaignVoiceBotSetting, CampaignVoiceUser
from EasyChatApp.models import Bot
from DeveloperConsoleApp.utils import get_developer_console_settings, send_email_to_customer_via_awsses
from CampaignApp.utils import logger

from datetime import datetime, timedelta
import json
import openpyxl
import os
import sys

from django.conf import settings
from cronjob_scripts.utils_campaign_cronjob_validator import check_if_cronjob_is_running, complete_cronjob_execution


def generate_report_excel(voice_campaign_history_export_request_obj, bot_obj, voice_campaign_user_objs):

    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Voice Campaign History Data"

    sheet.cell(1, 1).value = "Campaign Name"
    sheet.cell(1, 2).value = "Campaign SID"
    sheet.cell(1, 3).value = "Caller ID"
    sheet.cell(1, 4).value = "Call SID"
    sheet.cell(1, 5).value = "Status"
    sheet.cell(1, 6).value = "Date Created"
    sheet.cell(1, 7).value = "Type"
    sheet.cell(1, 8).value = "Disposition Code"
    sheet.cell(1, 9).value = "Total Duration"
    sheet.cell(1, 10).value = "On Call Duration"
    sheet.cell(1, 11).value = "Price"

    index = 1

    for voice_campaign_user_obj in voice_campaign_user_objs:
        sheet.cell(index + 1, 1).value = voice_campaign_user_obj.voice_campaign.campaign.name
        sheet.cell(index + 1, 2).value = voice_campaign_user_obj.voice_campaign.campaign_sid
        sheet.cell(index + 1, 3).value = voice_campaign_user_obj.mobile_number
        sheet.cell(index + 1, 4).value = voice_campaign_user_obj.get_call_sid()
        sheet.cell(index + 1, 5).value = voice_campaign_user_obj.get_call_status()
        sheet.cell(index + 1, 6).value = voice_campaign_user_obj.get_created_date_time()
        sheet.cell(index + 1, 7).value = "Outbound"
        sheet.cell(index + 1, 8).value = voice_campaign_user_obj.get_dispostion_code_data()[0]
        sheet.cell(index + 1, 9).value = voice_campaign_user_obj.get_call_duration()
        sheet.cell(index + 1, 10).value = str(voice_campaign_user_obj.on_call_duration)
        sheet.cell(index + 1, 11).value = str(voice_campaign_user_obj.price)

        index += 1

    if not os.path.exists(settings.MEDIA_ROOT + 'wpm_reports/bot-' + str(bot_obj.pk)):
        os.makedirs(settings.MEDIA_ROOT + 'wpm_reports/bot-' + str(bot_obj.pk))

    file_path = 'wpm_reports/bot-' + str(bot_obj.pk) + "/Voice_Campaign_History_Data_" + str(voice_campaign_history_export_request_obj.user.username) + ".xlsx"
    wb.save(settings.MEDIA_ROOT + file_path)
    return file_path


def cronjob():
    cronjob_detector_id = "campaign_export_cronjob"
    is_cronjob_exists, cronjob_tracker_obj = check_if_cronjob_is_running(cronjob_detector_id)
    if is_cronjob_exists:
        return
    try:
        voice_campaign_history_export_request_objs = VoiceCampaignHistoryExportRequest.objects.filter(is_completed=False)

        for voice_campaign_history_export_request_obj in voice_campaign_history_export_request_objs:
            try:
                request_data = json.loads(voice_campaign_history_export_request_obj.request_data)

                email_id = request_data["email_id"]
                campaign_ids = request_data["campaign_ids"]
                bot_pk = request_data["bot_pk"]
                filter_date_type = request_data["filter_date_type"]
                status_filter = request_data["status_filter"]

                bot_obj = Bot.objects.filter(pk=bot_pk, is_deleted=False).first()

                if not bot_obj:
                    continue

                start_date = None
                end_date = voice_campaign_history_export_request_obj.date_created

                filter_date_str = "beginning"

                if filter_date_type == "1":
                    start_date = voice_campaign_history_export_request_obj.date_created - timedelta(days=7)
                    filter_date_str = "{} to {}".format(start_date.strftime("%d-%m-%Y"), end_date.strftime("%d-%m-%Y"))
                elif filter_date_type == "2":
                    start_date = voice_campaign_history_export_request_obj.date_created - timedelta(days=30)
                    filter_date_str = "{} to {}".format(start_date.strftime("%d-%m-%Y"), end_date.strftime("%d-%m-%Y"))
                elif filter_date_type == "3":
                    start_date = voice_campaign_history_export_request_obj.date_created - timedelta(days=90)
                    filter_date_str = "{} to {}".format(start_date.strftime("%d-%m-%Y"), end_date.strftime("%d-%m-%Y"))
                elif filter_date_type == "5":
                    start_date = datetime.strptime(request_data["start_date"], "%Y-%m-%d").date()
                    end_date = datetime.strptime(request_data["end_date"], "%Y-%m-%d").date()
                    filter_date_str = "{} to {}".format(start_date.strftime("%d-%m-%Y"), end_date.strftime("%d-%m-%Y"))
                else:
                    end_date = None

                campaign_objs = Campaign.objects.filter(pk__in=campaign_ids)
                voice_campaign_objs = CampaignVoiceBotSetting.objects.filter(campaign__in=campaign_objs)
                voice_campaign_user_objs = CampaignVoiceUser.objects.filter(voice_campaign__in=voice_campaign_objs)

                if start_date:
                    voice_campaign_user_objs = voice_campaign_user_objs.filter(date__gte=start_date)
                if end_date:
                    voice_campaign_user_objs = voice_campaign_user_objs.filter(date__lte=end_date)

                if len(status_filter):
                    voice_campaign_user_objs = voice_campaign_user_objs.filter(status__in=status_filter)

                voice_campaign_user_objs = voice_campaign_user_objs.order_by("-pk")

                file_path = generate_report_excel(voice_campaign_history_export_request_obj, bot_obj, voice_campaign_user_objs)

                body = """
                    <head>
                        <meta http-equiv="Content-Type" content="text/html; charset=utf-8">
                        <title>Cogno AI</title>
                        <style type="text/css" media="screen">
                        </style>
                    </head>
                    <body>

                    <div style="padding:1em;border:0.1em black solid;" class="container">
                        <p>
                            Dear {},
                        </p>
                        <p>
                            We have received a request to provide you with the Voice Disposition report from {}. Please click on the link below to download the file.
                        </p>
                        <a href="{}/{}">click here</a>
                        <p>&nbsp;</p>"""

                config = get_developer_console_settings()

                body += config.custom_report_template_signature

                body += """</div></body>"""

                domain = settings.EASYCHAT_HOST_URL

                body = body.format(email_id, filter_date_str, domain, "files/" + file_path)

                send_email_to_customer_via_awsses(email_id, "Voice Disposition Report for {}".format(email_id), body)

                voice_campaign_history_export_request_obj.is_completed = True
                voice_campaign_history_export_request_obj.save()

            except Exception as e:
                exc_type, exc_obj, exc_tb = sys.exc_info()
                logger.error("cronjob: %s at %s", str(e), str(exc_tb.tb_lineno), extra={'AppName': 'Campaign'})

    except Exception as exc:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("cronjob: %s at %s", str(exc), str(exc_tb.tb_lineno), extra={'AppName': 'Campaign'})

    complete_cronjob_execution(cronjob_tracker_obj)
