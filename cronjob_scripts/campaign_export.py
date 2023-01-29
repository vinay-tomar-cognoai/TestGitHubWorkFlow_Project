from functools import reduce
import math
from CampaignApp.models import *
from DeveloperConsoleApp.utils import get_developer_console_settings, send_email_to_customer_via_awsses
import os
import sys
import pytz
import json

import logging
logger = logging.getLogger(__name__)

from django.conf import settings
from cronjob_scripts.utils_campaign_cronjob_validator import check_if_cronjob_is_running, complete_cronjob_execution
from CampaignApp.utils_external_api import get_mobile_number, get_audience_details_rcs, get_audience_log_unique_id_filtered_data, get_audience_log_phone_number_filtered_data, get_phone_number_list, add_audience_details_external_vb, operator
from CampaignApp.utils_export import *
from datetime import datetime, timedelta
import openpyxl
from os import path


def send_mail(from_email_id, to_emai_id, message_as_string, from_email_id_password):
    import smtplib
    # The actual sending of the e-mail
    server = smtplib.SMTP('smtp.gmail.com:587')
    # Credentials (if needed) for sending the mail
    password = from_email_id_password
    # Start tls handshaking
    server.starttls()
    # Login to server
    server.login(from_email_id, password)
    # Send mail
    server.sendmail(from_email_id, to_emai_id, message_as_string)
    # Close session
    server.quit()


def add_default_columns_vb(sheet):
    sheet.cell(1, 1).value = "Campaign Name"
    sheet.cell(1, 2).value = "Audience Batch Name"
    sheet.cell(1, 3).value = "Total Audience Batch Size"
    sheet.cell(1, 4).value = "Call scheduled"
    sheet.cell(1, 5).value = "Call Initiated"
    sheet.cell(1, 6).value = "Call completed"
    sheet.cell(1, 7).value = "Failed"
    sheet.cell(1, 8).value = "Inprogress"
    sheet.cell(1, 9).value = "Invalid"
    sheet.cell(1, 10).value = "App ID"


def add_default_columns_rcs(sheet):
    sheet.cell(1, 1).value = "Campaign Name"
    sheet.cell(1, 2).value = "Audience Batch Name"
    sheet.cell(1, 3).value = "Total Audience Batch Size"
    sheet.cell(1, 4).value = "Message Sent"
    sheet.cell(1, 5).value = "Message Delivered"
    sheet.cell(1, 6).value = "Message Read"
    sheet.cell(1, 7).value = "Failed"
    sheet.cell(1, 8).value = "Template Name"


def add_campaign_details_vb(sheet, campaign_obj):
    voice_bot_obj = CampaignVoiceBotSetting.objects.filter(
        campaign=campaign_obj).first()
    campaign_analytics = CampaignVoiceBotAnalytics.objects.filter(
        campaign=campaign_obj).first()

    if campaign_obj.batch:
        sheet.cell(1, 1).value = campaign_obj.name
        sheet.cell(1, 2).value = campaign_obj.batch.batch_name
        sheet.cell(1, 3).value = campaign_obj.batch.total_audience
        sheet.cell(1, 4).value = campaign_analytics.call_scheduled
        sheet.cell(1, 5).value = campaign_analytics.call_initiated
        sheet.cell(1, 6).value = campaign_analytics.call_completed
        sheet.cell(1, 7).value = campaign_analytics.call_failed
        sheet.cell(1, 8).value = campaign_analytics.call_in_progress
        sheet.cell(1, 9).value = campaign_analytics.call_invalid
        sheet.cell(1, 10).value = voice_bot_obj.app_id


def add_campaign_details_rcs(sheet, campaign_obj):
    rcs_analytics_obj = CampaignRCSDetailedAnalytics.objects.filter(
        campaign=campaign_obj).first()

    if campaign_obj.batch and campaign_obj.campaign_template_rcs:
        sheet.cell(2, 1).value = campaign_obj.name
        sheet.cell(2, 2).value = campaign_obj.batch.batch_name
        sheet.cell(2, 3).value = rcs_analytics_obj.submitted
        sheet.cell(2, 4).value = rcs_analytics_obj.sent
        sheet.cell(2, 5).value = rcs_analytics_obj.delivered
        sheet.cell(2, 6).value = rcs_analytics_obj.read
        sheet.cell(2, 7).value = rcs_analytics_obj.failed
        sheet.cell(2, 8).value = campaign_obj.campaign_template_rcs.template_name


def add_default_detailed_columns_vb(sheet):

    sheet.cell(1, 1).value = "Campaign Name"
    sheet.cell(1, 2).value = "Campaign SID"
    sheet.cell(1, 3).value = "Caller ID"
    sheet.cell(1, 4).value = "Call SID"
    sheet.cell(1, 5).value = "Status"
    sheet.cell(1, 6).value = "Date Created"
    sheet.cell(1, 7).value = "Type"
    sheet.cell(1, 8).value = "Disposition Code"
    sheet.cell(1, 9).value = "Total Duration"
    sheet.cell(1, 10).value = "On Call Duration"
    sheet.cell(1, 11).value = "Price"


def add_default_external_detailed_columns_vb(sheet, detail_report_headers):

    for idx, header in enumerate(detail_report_headers):
        sheet.cell(1, idx + 1).value = header


def add_default_detailed_columns_rcs(sheet, is_filter_on_export):
    sheet.cell(1, 1).value = "Mobile Number"
    sheet.cell(1, 2).value = "Recipient ID"
    sheet.cell(1, 3).value = "Sent Time"
    sheet.cell(1, 4).value = "Delivered Time"
    sheet.cell(1, 5).value = "Read Time"
    sheet.cell(1, 6).value = "Template Type"
    sheet.cell(1, 7).value = "Template Name"
    sheet.cell(1, 8).value = "Status Code"
    if is_filter_on_export:
        sheet.cell(1, 9).value = "Unique ID"


def add_audience_details_rcs(audience_obj, audience_log_obj, campaign_obj, row, sheet, masking_enabled, is_filter_on_export):
    try:
        mobile_number = get_mobile_number(
            audience_obj.audience_id, False, masking_enabled)
        audience_details_rcs = get_audience_details_rcs(
            audience_log_obj, campaign_obj)

        sheet.cell(2, 1).value = mobile_number
        sheet.cell(2, 2).value = audience_details_rcs["Recipient ID"]
        sheet.cell(2, 3).value = audience_details_rcs["Sent Time"]
        sheet.cell(2, 4).value = audience_details_rcs["Delivered Time"]
        sheet.cell(2, 5).value = audience_details_rcs["Read Time"]
        sheet.cell(2, 6).value = audience_details_rcs["Template Type"]
        sheet.cell(2, 7).value = audience_details_rcs["Template Name"]
        sheet.cell(2, 8).value = audience_details_rcs["Status Code"]
        if is_filter_on_export:
            audience_unique_id = audience_obj.audience_unique_id
            sheet.cell(2, 9).value = audience_unique_id if audience_unique_id else "-"

    except Exception as exc:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("add_audience_details_rcs: %s at %s", str(exc), str(
            exc_tb.tb_lineno), extra={'AppName': 'Campaign'})


def add_audience_details_vb(voice_campaign_user_obj, row, sheet, masking_enabled):
    try:
        # making it false because masking is currently not needed in voice bot
        masking_enabled = False
        sheet.cell(row + 1, 1).value = voice_campaign_user_obj.voice_campaign.campaign.name
        sheet.cell(row + 1, 2).value = voice_campaign_user_obj.voice_campaign.campaign_sid
        mobile_number = get_mobile_number(
            voice_campaign_user_obj.mobile_number, True, masking_enabled)
        sheet.cell(row + 1, 3).value = mobile_number
        sheet.cell(row + 1, 4).value = voice_campaign_user_obj.get_call_sid()
        sheet.cell(row + 1, 5).value = voice_campaign_user_obj.get_call_status()
        sheet.cell(row + 1, 6).value = voice_campaign_user_obj.get_created_date_time()
        sheet.cell(row + 1, 7).value = "Outbound"
        sheet.cell(row + 1, 8).value = voice_campaign_user_obj.get_dispostion_code_data()[0]
        sheet.cell(row + 1, 9).value = voice_campaign_user_obj.get_call_duration()
        sheet.cell(row + 1, 10).value = str(voice_campaign_user_obj.on_call_duration)
        sheet.cell(row + 1, 11).value = str(voice_campaign_user_obj.price)

    except Exception as exc:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("add_audience_details_vb: %s at %s", str(exc), str(
            exc_tb.tb_lineno), extra={'AppName': 'Campaign'})


def add_external_api_audience_details_vb(voice_campaign_user_data, row, sheet):
    try:

        for idx, user_data in enumerate(voice_campaign_user_data):
            sheet.cell(row + 1, idx + 1).value = user_data

    except Exception as exc:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("add_audience_details_vb: %s at %s", str(exc), str(
            exc_tb.tb_lineno), extra={'AppName': 'Campaign'})


def generate_external_api_report_vb(test_wb, campaign_obj, filters_on_export):
    try:
        if not os.path.exists(settings.MEDIA_ROOT + 'wpm_reports/bot-' + str(campaign_obj.bot.pk)):
            os.makedirs(settings.MEDIA_ROOT + 'wpm_reports/bot-' +
                        str(campaign_obj.bot.pk))

        test_wb = openpyxl.Workbook()

        sheet1 = test_wb['Sheet']
        sheet1.title = "Overall Report"

        sheet2 = test_wb.create_sheet("Detailed Report")

        add_default_columns_vb(sheet1)
        add_campaign_details_vb(sheet1, campaign_obj)

        call_details_objs = CampaignVoiceBotDetailedAnalytics.objects.filter(
            campaign=campaign_obj)
        phone_number = filters_on_export['phone_number']
        audience_unique_id = filters_on_export['audience_unique_id']
        if phone_number:
            phone_number = get_phone_number_list(phone_number)
            if phone_number:
                query = reduce(operator.or_, (Q(from_number=number) for number in phone_number))
                call_details_objs = call_details_objs.filter(query)
        
        if audience_unique_id:
            call_details_objs = get_audience_log_unique_id_filtered_data(audience_unique_id, call_details_objs)  
        is_sheet2_header_written = False

        row = 1
        for call_details_obj in call_details_objs.iterator():
            response_detailed_json_obj = add_audience_details_external_vb(call_details_obj, campaign_obj)

            if not is_sheet2_header_written:
                add_default_external_detailed_columns_vb(sheet2, response_detailed_json_obj.keys())
                is_sheet2_header_written = True

            add_external_api_audience_details_vb(
                response_detailed_json_obj.values(), row, sheet2)
            row += 1

        filename = 'wpm_reports/bot-' + str(campaign_obj.bot.pk) + \
            '/' + campaign_obj.name.replace(' ', '') + \
            '_' + str(datetime.now()) + '.xlsx'

        test_wb.save(settings.MEDIA_ROOT + filename)

        return filename
    except Exception as exc:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("generate_report_vb: %s at %s", str(exc), str(
            exc_tb.tb_lineno), extra={'AppName': 'Campaign'})


def generate_test_report(campaign_obj, masking_enabled):
    filename = []
    try:
        audience_log_objs = CampaignAudienceLog.objects.filter(campaign=campaign_obj, is_test=True).values('id', 'audience__audience_id', 'audience__record', 'recepient_id',
                                                                                                           'is_sent', 'sent_datetime', 'is_delivered', 'delivered_datetime', 'is_read', 'read_datetime', 'is_replied', 'replied_datetime', 'is_failed', 'request', 'response')

        if not os.path.exists(settings.MEDIA_ROOT + 'wpm_reports/bot-' + str(campaign_obj.bot.pk)):
            os.makedirs(settings.MEDIA_ROOT + 'wpm_reports/bot-' +
                        str(campaign_obj.bot.pk))

        campaign_details = get_test_campaign_details_data(campaign_obj)
        whatsapp_bsp_name = campaign_details["whatsapp_bsp_name"]
        if whatsapp_bsp_name == "-":
            whatsapp_bsp_name = "Ameyo"

        sheet_count = 1
        while (sheet_count <= 1):  # Hardcode sheet max value is 1 for test reports
            test_wb = openpyxl.Workbook(write_only=True)
            sheet1 = test_wb.create_sheet("Overall Test report")
            sheet2 = test_wb.create_sheet("Detailed Test report")
            add_campaign_test_details(sheet1, campaign_details)
            add_whatsapp_test_sheet_columns(sheet2)

            for audience_log in audience_log_objs.iterator():
                add_test_audience_details(
                    audience_log, sheet2, masking_enabled, whatsapp_bsp_name)

            filepath = 'wpm_reports/bot-' + str(campaign_obj.bot.pk) + \
                '/' + campaign_details["campaign_name"].replace(' ', '') + \
                '_test_report' + '_' + str(datetime.now()) + '.xlsx'

            test_wb.save(settings.MEDIA_ROOT + filepath)
            filename.append(filepath)
            sheet_count += 1

    except Exception as exc:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("generate_test_report: %s at %s", str(exc), str(
            exc_tb.tb_lineno), extra={'AppName': 'Campaign'})

    return filename


def generate_report(campaign_obj, masking_enabled, filters_on_export=dict()):
    filename = []
    try:
        if not os.path.exists(settings.MEDIA_ROOT + 'wpm_reports/bot-' + str(campaign_obj.bot.pk)):
            os.makedirs(settings.MEDIA_ROOT + 'wpm_reports/bot-' +
                        str(campaign_obj.bot.pk))

        phone_number = filters_on_export.get('phone_number')
        audience_unique_id = filters_on_export.get('audience_unique_id')

        audience_log_objs = CampaignAudienceLog.objects.filter(
            campaign=campaign_obj, is_test=False)

        if phone_number:
            phone_number = get_phone_number_list(phone_number)
            audience_log_objs = get_audience_log_phone_number_filtered_data(
                phone_number, audience_log_objs)
        if audience_unique_id:
            audience_log_objs = get_audience_log_unique_id_filtered_data(
                audience_unique_id, audience_log_objs)

        audience_log_objs = audience_log_objs.values('id', 'audience__audience_unique_id', 'audience__audience_id', 'audience__record', 'recepient_id', 'is_sent',
                                                     'sent_datetime', 'is_delivered', 'delivered_datetime', 'is_read', 'read_datetime', 'is_replied', 'replied_datetime', 'is_failed', 'request', 'response')

        total_audience = audience_log_objs.count()
        total_sheet = math.ceil(total_audience / CAMPAIGN_REPORT_MAX_ROWS)
        if not total_sheet:
            return []

        campaign_details = get_campaign_details_data(campaign_obj)
        whatsapp_bsp_name = campaign_details["whatsapp_bsp_name"]
        if whatsapp_bsp_name == "-":
            whatsapp_bsp_name = "Ameyo"

        is_external_campaign_export = False if campaign_details["source"] == "Dashboard" else True

        sheet_count = 1
        start_audience_index = 0
        while (sheet_count <= total_sheet):
            test_wb = openpyxl.Workbook(write_only=True)
            sheet1 = test_wb.create_sheet("Overall report")
            sheet2 = test_wb.create_sheet("Detailed report")
            add_campaign_details(sheet1, campaign_details)
            add_default_detailed_columns(sheet2, is_external_campaign_export)

            for audience_log in audience_log_objs[start_audience_index:CAMPAIGN_REPORT_MAX_ROWS * sheet_count].iterator():
                add_audience_details(
                    audience_log, sheet2, masking_enabled, whatsapp_bsp_name, is_external_campaign_export)

            filepath = 'wpm_reports/bot-' + str(campaign_obj.bot.pk) + \
                '/' + campaign_details["campaign_name"].replace(' ', '') + \
                '_Part-' + str(sheet_count) + '_' + \
                str(datetime.now()) + '.xlsx'

            test_wb.save(settings.MEDIA_ROOT + filepath)
            filename.append(filepath)
            start_audience_index = CAMPAIGN_REPORT_MAX_ROWS * sheet_count
            sheet_count += 1

    except Exception as exc:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("generate_report: %s at %s", str(exc), str(
            exc_tb.tb_lineno), extra={'AppName': 'Campaign'})

    return filename


def generate_report_single_campaign_date_range(test_wb, campaign_obj, masking_enabled, start_date, end_date, filters_on_export):
    filename = []
    try:
        if not os.path.exists(settings.MEDIA_ROOT + 'wpm_reports/bot-' + str(campaign_obj.bot.pk)):
            os.makedirs(settings.MEDIA_ROOT + 'wpm_reports/bot-' +
                        str(campaign_obj.bot.pk))

        phone_number = filters_on_export.get('phone_number')
        audience_unique_id = filters_on_export.get('audience_unique_id')

        audience_log_objs = CampaignAudienceLog.objects.filter(
            campaign=campaign_obj, is_test=False)

        if phone_number:
            phone_number = get_phone_number_list(phone_number)
            audience_log_objs = get_audience_log_phone_number_filtered_data(
                phone_number, audience_log_objs)
        if audience_unique_id:
            audience_log_objs = get_audience_log_unique_id_filtered_data(
                audience_unique_id, audience_log_objs)

        audience_log_objs = audience_log_objs.values(
            'audience__audience_unique_id', 'audience__audience_id', 'audience__record', 'recepient_id', 'is_sent',
            'sent_datetime', 'is_delivered', 'delivered_datetime', 'is_read', 'read_datetime', 'is_replied', 'replied_datetime',
            'is_failed', 'request', 'response', 'quick_replies__name', 'id')

        total_audience = audience_log_objs.count()
        total_sheet = math.ceil(total_audience / CAMPAIGN_REPORT_MAX_ROWS)
        if not total_sheet:
            return []

        campaign_details = get_campaign_details_data(campaign_obj)
        whatsapp_bsp_name = campaign_details["whatsapp_bsp_name"]
        if whatsapp_bsp_name == "-":
            whatsapp_bsp_name = "Ameyo"

        is_external_campaign_export = False if campaign_details["source"] == "Dashboard" else True

        sheet_count = 1
        start_audience_index = 0
        while (sheet_count <= total_sheet):
            test_wb = openpyxl.Workbook(write_only=True)
            sheet1 = test_wb.create_sheet("Overall report")
            sheet2 = test_wb.create_sheet("Detailed report")
            add_campaign_details(sheet1, campaign_details)
            add_default_detailed_columns(sheet2, is_external_campaign_export)

            for audience_log in audience_log_objs[start_audience_index:CAMPAIGN_REPORT_MAX_ROWS * sheet_count].iterator():
                add_audience_details(
                    audience_log, sheet2, masking_enabled, whatsapp_bsp_name, is_external_campaign_export)

            filepath = 'wpm_reports/bot-' + str(campaign_obj.bot.pk) + \
                '/' + campaign_details["campaign_name"].replace(' ', '') + \
                "_from_" + start_date.strftime("%d-%m-%Y") + \
                "_to_" + end_date.strftime("%d-%m-%Y") + \
                '_' + str(datetime.now()) + '.xlsx'

            test_wb.save(settings.MEDIA_ROOT + filepath)
            filename.append(filepath)
            start_audience_index = CAMPAIGN_REPORT_MAX_ROWS * sheet_count
            sheet_count += 1

    except Exception as exc:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("generate_report_single_campaign_date_range: %s at %s", str(exc), str(
            exc_tb.tb_lineno), extra={'AppName': 'Campaign'})

    return filename


def generate_report_vb(test_wb, campaign_obj, masking_enabled):
    filename = None
    try:
        if not os.path.exists(settings.MEDIA_ROOT + 'wpm_reports/bot-' + str(campaign_obj.bot.pk)):
            os.makedirs(settings.MEDIA_ROOT + 'wpm_reports/bot-' +
                        str(campaign_obj.bot.pk))
        sheet1 = test_wb['Sheet']
        sheet1.title = "Overall Report"
        sheet1 = test_wb['Sheet']
        sheet1.title = "Overall Report"
        sheet2 = test_wb.create_sheet("Detailed Report")

        add_default_columns_vb(sheet1)
        add_campaign_details_vb(sheet1, campaign_obj)
        add_default_detailed_columns_vb(sheet2)

        voice_campaign_obj = CampaignVoiceBotSetting.objects.filter(campaign=campaign_obj).first()
        voice_campaign_user_objs = CampaignVoiceUser.objects.filter(voice_campaign=voice_campaign_obj)

        row = 1
        for voice_campaign_user_obj in voice_campaign_user_objs:

            add_audience_details_vb(
                voice_campaign_user_obj, row, sheet2, masking_enabled)
            row += 1

        filename = 'wpm_reports/bot-' + str(campaign_obj.bot.pk) + \
            '/' + campaign_obj.name.replace(' ', '') + \
            '_' + str(datetime.now()) + '.xlsx'

        test_wb.save(settings.MEDIA_ROOT + filename)

    except Exception as exc:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("generate_report_vb: %s at %s", str(exc), str(
            exc_tb.tb_lineno), extra={'AppName': 'Campaign'})

    return filename


def generate_report_rcs(test_wb, campaign_obj, masking_enabled, filters_on_export=dict()):
    filename = None
    try:
        if not os.path.exists(settings.MEDIA_ROOT + 'wpm_reports/bot-' + str(campaign_obj.bot.pk)):
            os.makedirs(settings.MEDIA_ROOT + 'wpm_reports/bot-' +
                        str(campaign_obj.bot.pk))
        sheet1 = test_wb['Sheet']
        sheet1.title = "Overall Report"

        sheet2 = test_wb.create_sheet("Detailed Report")

        is_filter_on_export = True if filters_on_export else False
        
        add_default_columns_rcs(sheet1)
        add_campaign_details_rcs(sheet1, campaign_obj)
        add_default_detailed_columns_rcs(sheet2, is_filter_on_export)

        audience_objs = CampaignAudience.objects.filter(campaign=campaign_obj)
        
        phone_number = filters_on_export.get('phone_number')
        audience_unique_id = filters_on_export.get('audience_unique_id')
        if phone_number:
            phone_number = get_phone_number_list(phone_number)
            audience_objs = get_audience_log_phone_number_filtered_data(phone_number, audience_objs)
        if audience_unique_id:
            audience_objs = get_audience_log_unique_id_filtered_data(audience_unique_id, audience_objs)

        row = 1
        for audience_obj in audience_objs:
            audience_log = CampaignAudienceLog.objects.filter(
                audience=audience_obj, campaign=campaign_obj).first()
            if audience_log:
                add_audience_details_rcs(
                    audience_obj, audience_log, campaign_obj, row, sheet2, masking_enabled, is_filter_on_export)
                row += 1

        filename = 'wpm_reports/bot-' + str(campaign_obj.bot.pk) + \
            '/' + campaign_obj.name.replace(' ', '') + \
            '_' + str(datetime.now()) + '.xlsx'

        test_wb.save(settings.MEDIA_ROOT + filename)

    except Exception as exc:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("generate_report_rcs: %s at %s", str(exc), str(
            exc_tb.tb_lineno), extra={'AppName': 'Campaign'})

    return filename


def get_email_body(email_id, file_path, format_data):

    body_sentence = format_data.get('body_sentence', '')
    details = format_data.get('details', '')
    domain = settings.EASYCHAT_HOST_URL
    
    return f"""
            <head>
                <meta http-equiv="Content-Type" content="text/html; charset=utf-8">
                <title>Cogno AI</title>
                <style type="text/css" media="screen">
                </style>
            </head>
            <body>

            <div style="padding:1em;border:0.1em black solid;" class="container">
                <p>
                    Dear {email_id},
                </p>
                <p>
                    {body_sentence}<br>
                    Please click on the link to download the file.
                    <a href="{domain}/{file_path}">click here</a><br><br>
                    {details}
                </p>
                <p>
                    
                </p>
                
                <p>&nbsp;</p>"""


def cronjob():
    cronjob_detector_id = "campaign_export_cronjob"
    is_cronjob_exists, cronjob_tracker_obj = check_if_cronjob_is_running(
        cronjob_detector_id)
    if is_cronjob_exists:
        return
    time_zone = pytz.timezone(settings.TIME_ZONE)
    try:
        message_history_datadump_log = open(
            "log/campaign_export.log", "a")

        if not os.path.exists(settings.MEDIA_ROOT + 'wpm_reports'):
            os.makedirs(settings.MEDIA_ROOT + 'wpm_reports')

        export_requests = CampaignExportRequest.objects.filter(
            is_completed=False)

        for export_request in export_requests.iterator():
            masking_enabled = export_request.masking_enabled
            filters_on_export = json.loads(export_request.filters_on_export)

            if export_request.export_type == '1':
                campaign_obj = export_request.campaign

                test_wb = openpyxl.Workbook()

                if campaign_obj.channel.name == 'Voice Bot':
                    if filters_on_export:
                        file_path = generate_external_api_report_vb(
                            test_wb, campaign_obj, filters_on_export)
                        file_path = [file_path] if file_path else []
                    else:
                        file_path = generate_report_vb(
                            test_wb, campaign_obj, masking_enabled)
                        file_path = [file_path] if file_path else []
                elif campaign_obj.channel.name == 'RCS':
                    file_path = generate_report_rcs(
                        test_wb, campaign_obj, masking_enabled, filters_on_export)
                    file_path = [file_path] if file_path else []
                else:
                    file_path = generate_report(
                        campaign_obj, masking_enabled, filters_on_export)

                if not file_path:
                    logger.error("campaign_export cronjob: filepath not found for request for export_request with id %s", str(
                        export_request.pk), extra={'AppName': 'Campaign'})
                    continue

                if len(file_path) == 1:
                    file_path = 'files/' + file_path[0]
                else:
                    export_zip_file_path = "files/wpm_reports/bot-" + \
                        str(export_request.bot.pk) + "/CampaignReport-" + \
                        campaign_obj.name + "_" + str(datetime.now()) + ".zip"

                    file_path = get_zip_file_path(
                        file_path, export_zip_file_path)

                body_sentence = 'We have received a request to provide you with the Campaign report'
            elif export_request.export_type == '2':
                start_date = export_request.start_date.astimezone(time_zone)
                end_date = export_request.end_date.astimezone(time_zone)

                if not os.path.exists(settings.MEDIA_ROOT + 'wpm_reports/bot-' + str(export_request.bot.pk)):
                    os.makedirs(settings.MEDIA_ROOT + 'wpm_reports/bot-' +
                                str(export_request.bot.pk))

                campaign_objs = Campaign.objects.filter(
                    bot=export_request.bot, create_datetime__date__gte=start_date, create_datetime__date__lte=end_date, is_deleted=False, channel__is_deleted=False)

                file_paths = []

                for campaign_obj in campaign_objs.iterator():
                    test_wb = openpyxl.Workbook()

                    if campaign_obj.channel.name == 'Voice Bot':
                        file_path = generate_report_vb(
                            test_wb, campaign_obj, masking_enabled)
                    elif campaign_obj.channel.name == 'RCS':
                        file_path = generate_report_rcs(
                            test_wb, campaign_obj, masking_enabled)
                    else:
                        file_path = generate_report(
                            campaign_obj, masking_enabled)

                    if file_path:
                        file_paths.extend(file_path)

                if not file_paths:
                    logger.error("campaign_export cronjob: filepath not found for request for export_request with id %s", str(
                        export_request.pk), extra={'AppName': 'Campaign'})
                    continue

                export_zip_file_path = "files/wpm_reports/bot-" + \
                    str(export_request.bot.pk) + "/CampaignReportCustom-" + \
                    str(export_request.pk) + "_from_" + start_date.strftime("%d-%m-%Y") + \
                    "_to_" + end_date.strftime("%d-%m-%Y") + ".zip"

                file_path = get_zip_file_path(file_paths, export_zip_file_path)

                body_sentence = 'We have received a request to provide you with the Campaign report'

            else:
                # this export type is only for whatsapp event based api for now
                campaign_obj = export_request.campaign
                test_wb = openpyxl.Workbook()

                start_date = export_request.start_date.astimezone(time_zone)
                end_date = export_request.end_date.astimezone(time_zone)

                file_path = generate_report_single_campaign_date_range(
                    test_wb, campaign_obj, masking_enabled, start_date, end_date, filters_on_export)
                if not file_path:
                    logger.error("campaign_export cronjob: filepath not found for request for export_request with id %s", str(
                        export_request.pk), extra={'AppName': 'Campaign'})
                    continue

                if len(file_path) == 1:
                    file_path = 'files/' + file_path[0]
                else:
                    export_zip_file_path = "files/wpm_reports/bot-" + \
                        str(export_request.bot.pk) + "/CampaignReport-" + \
                        campaign_obj.name + "_" + str(datetime.now()) + ".zip"

                    file_path = get_zip_file_path(
                        file_path, export_zip_file_path)

                body_sentence = 'We have received a request to provide you with the Campaign report'

            email_id = export_request.email_id
            bot_name = export_request.bot.name

            format_data = {
                "body_sentence": body_sentence
            }
            
            if export_request.export_type == '1':
                campaign_name = export_request.campaign.name
                template_obj = export_request.campaign.campaign_template
                template_name = template_obj.template_name if template_obj else '-'
                details = f"""<b>Campaign Name:</b> {campaign_name}<br>
                    <b>Bot Name:</b> {bot_name}<br>
                    <b>Template Name:</b> {template_name}<br>
                    """
                format_data["details"] = details
                body = get_email_body(email_id, file_path, format_data)
            else:
                start_date = datetime.strftime(start_date, "%d-%m-%Y")
                end_date = datetime.strftime(end_date, "%d-%m-%Y")
                details = f"""
                    <b>Bot Name:</b> {bot_name}<br>
                    <b>Date Range:</b> {start_date} - {end_date}<br>
                    """
                format_data["details"] = details
                body = get_email_body(email_id, file_path, format_data)

            config = get_developer_console_settings()

            body += config.custom_report_template_signature

            body += """</div></body>"""

            send_email_to_customer_via_awsses(
                email_id, f"Campaign Report For {bot_name} - {email_id}", body)

            export_request.is_completed = True
            export_request.save()

        message_history_datadump_log.close()
    except Exception as exc:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("cronjob: %s at %s", str(exc), str(
            exc_tb.tb_lineno), extra={'AppName': 'Campaign'})
    complete_cronjob_execution(cronjob_tracker_obj)
