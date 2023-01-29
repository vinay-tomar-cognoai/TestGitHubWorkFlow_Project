import os
import base64
from os import error
from Crypto.Cipher import AES
from django.shortcuts import render, redirect, HttpResponse
from django.core.cache import cache
# Django REST framework
import requests
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView
from django.utils.encoding import smart_str
from rest_framework.authentication import SessionAuthentication, \
    BasicAuthentication
from django.core.paginator import Paginator
from django.contrib.auth.decorators import login_required
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from django import forms

from CampaignApp.html_parser import strip_html_tags
from CampaignApp.constants import *
from CampaignApp.utils_client_server_signal import send_data_from_server_to_client
from cronjob_scripts.utils_campaign_cronjob_validator import *
from CampaignApp.utils_aws_sqs import send_message_into_campaign_sqs

import json
from django.utils.dateformat import DateFormat
from django.utils.formats import get_format
from CampaignApp.utils_custom_encryption import *
from CampaignApp.utils_parse import *
from CampaignApp.utils_campaign_rcs import *
from CampaignApp.utils_validation import CampaignInputValidation
from CampaignApp.utils_export import add_sheet_whatsapp_audience_history_report, add_status_for_bsp, add_default_whatsapp_audience_history_columns, add_whatsapp_overall_details, get_campaign_details_data
from CampaignApp.models import *
from EasyChatApp.models import Profile, RCSDetails, WhatsappCredentialsConfig
from django.db.models import Q, Count, Sum, F
from dateutil import relativedelta
from EasyChatApp.rcs_business_messaging import rbm_service, messages
import operator
import openpyxl
from functools import reduce
import boto3


import uuid
import calendar
import threading
import sys
import re
from django.conf import settings
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import pytz
import magic
import xlrd
import func_timeout
import mimetypes
import math
from xlsxwriter import Workbook
from datetime import datetime, timedelta
import xlwt
import csv


from DeveloperConsoleApp.utils import send_email_to_customer_via_awsses

# Logger
import logging
logger = logging.getLogger(__name__)
validation_obj = CampaignInputValidation()


def logout_all(username, UserSession, Session):
    try:
        logger.info("In logout_all", extra={
                    'AppName': 'Campaign', 'user_id': str(username)})
        user_session_objs = UserSession.objects.filter(user__username=username)
        for user_session_obj in user_session_objs:
            delete_user_session(user_session_obj, Session)
    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("In logout_all: %s at %s", str(e), str(exc_tb.tb_lineno), extra={
                     'AppName': 'Campaign', 'user_id': str(username)})


def delete_user_session(user_session, Session):
    try:
        logger.info("In delete_user_session", extra={'AppName': 'Campaign',
                                                     'user_id': user_session.user.username})
        session_objs = Session.objects.filter(pk=user_session.session_key)
        user_session.delete()
        if session_objs:
            session_objs.delete()
    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("In delete_user_session: %s at %s", str(e), str(exc_tb.tb_lineno), extra={
            'AppName': 'Campaign', 'user_id': str(user_session.user.username)})


"""

This function returns True if the input email is valid

"""


def check_valid_email(email):
    regex = "(^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$)"
    if(re.search(regex, email)):
        return True
    return False


def send_otp_mail(email_id, otp, session_time):

    body = """
    <head>
      <meta http-equiv="Content-Type" content="text/html; charset=utf-8">
      <title>Cogno AI</title>
      <style type="text/css" media="screen">
      </style>
    </head>
    <body>

    <div style="padding:1em;border:0.1em black solid;" class="container">
        <div style="width: fit-content">
        <p>
            Hello, <br><br>Please use the OTP code below to verify yourself for campaign API integration. Your session will expire in """ + session_time + """ hours.
        </p>
        <div style="display: block; width: max-content; margin: auto;"><p style="word-break: break-word; color: red; padding: 3px 10px; font-weight: bold; font-size: 18px;">""" + otp + """</p></div>
        </div>
        <p>
            Thanks,<br>CognoAI Team
        <p>
    </div>

    </body>"""

    send_email_to_customer_via_awsses(
        email_id, "OTP Code to access API Integration", body)


"""
function: send_mail
input params:
    from_email_id: Sender email id
    to_email_id: Receiver email id
    message_as_string: Email message string
    from_email_id_password: Sender email password

Send mail to given email id
"""


# def send_mail(from_email_id, to_emai_id, message_as_string, from_email_id_password):
#     import smtplib
#     # The actual sending of the e-mail
#     server = smtplib.SMTP('smtp.gmail.com:587')
#     # Credentials (if needed) for sending the mail
#     password = from_email_id_password
#     # Start tls handshaking
#     server.starttls()
#     # Login to server
#     server.login(from_email_id, password)
#     # Send mail
#     server.sendmail(from_email_id, to_emai_id, message_as_string)
#     # Close session
#     server.quit()


def open_file(file_dir, method):
    try:
        file_dir = settings.SECURE_MEDIA_ROOT + file_dir

        if '..' in file_dir:
            logger.error("user is trying to access this file: %s", str(file_dir), extra={
                'AppName': 'Campaign'})
            return

    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("open_file: %s %s", str(e), str(exc_tb.tb_lineno), extra={
            'AppName': 'Campaign'})

    return open(file_dir, method)


def get_config_obj(CampaignConfig):
    try:
        config_obj = CampaignConfig.objects.all()

        if config_obj:
            return config_obj[0]

        config_obj = CampaignConfig.objects.create()

        return config_obj
    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("get_config_obj: %s %s", str(e), str(exc_tb.tb_lineno), extra={
            'AppName': 'Campaign'})

        config_obj = CampaignConfig.objects.create()

        return config_obj


def remo_html_from_string(raw_str):
    try:
        regex_cleaner = re.compile('<.*?>')
        cleaned_raw_str = re.sub(regex_cleaner, '', raw_str)
        return cleaned_raw_str
    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("In remo_html_from_string: %s at %s", str(e), str(exc_tb.tb_lineno), extra={
            'AppName': 'Campaign'})
        return raw_str


def remo_special_tag_from_string(raw_str):
    try:
        cleaned_raw_str = raw_str.replace(
            "+", "").replace("|", "").replace("-", "").replace("=", "").replace("<", "").replace(">", "")
        return cleaned_raw_str
    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("In remo_special_tag_from_string: %s at %s", str(e), str(exc_tb.tb_lineno), extra={
            'AppName': 'Campaign'})
        return raw_str


def remo_special_tag_from_batch_name(raw_str):
    try:
        cleaned_raw_str = raw_str.replace(
            "+", "").replace("|", "").replace("=", "").replace("<", "").replace(">", "")
        return cleaned_raw_str
    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("In remo_special_tag_from_string: %s at %s", str(e), str(exc_tb.tb_lineno), extra={
            'AppName': 'Campaign'})
        return raw_str


def clean_input(str):
    try:
        str = str.strip()
        str = remo_html_from_string(str)
        str = remo_special_tag_from_string(str)

    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("In clean_input: %s at %s", str(e), str(exc_tb.tb_lineno), extra={
            'AppName': 'Campaign'})

    return str


def remove_special_chars_from_filename(filename):
    try:
        cleaned_filename = filename.replace(
            "&", "").replace("@", "").replace("<", "").replace(">", "")
        return cleaned_filename
    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("In remove_special_chars_from_filename: %s at %s", str(e), str(exc_tb.tb_lineno), extra={
            'AppName': 'Campaign'})
        return filename


def check_malicious_file_from_filename(filename, allowed_files_list=None):
    if allowed_files_list == None:
        allowed_files_list = [
            "png", "jpg", "jpeg", "bmp", "gif", "tiff", "exif", "jfif", "webm", "mpg", "jpe",
            "mp2", "mpeg", "mpe", "mpv", "ogg", "mp4", "m4p", "m4v", "avi", "wmv", "mov", "qt", "doc", "docx",
            "flv", "swf", "avchd", "mp3", "aac", "pdf", "xls", "xlsx", "json", "xlsm", "xlt", "xltm", "zip", "xlb",
            "bin",
        ]

    try:
        if len(filename.split('.')) != 2:
            return True

        file_extension = filename.split('.')[-1].lower()

        if file_extension not in allowed_files_list:
            return True
        else:
            return False
    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("In check_malicious_file_from_filename: %s at %s", str(e), str(exc_tb.tb_lineno), extra={
            'AppName': 'Campaign'})
        return True


def check_malicious_file(uploaded_file):
    return check_malicious_file_from_filename(uploaded_file.name)


def check_malicious_file_from_content(base64_data, allowed_files_list=None):
    try:

        # Untill we find any concrete solution for detecting file type
        return False

        # decoded = base64.b64decode(base64_data)
        # mime_type = magic.from_buffer(decoded[:10000], mime=True)
        # file_ext = mimetypes.guess_extension(mime_type)

        # logger.info("In check_malicious_file_from_content file_ext is: %s ", str(file_ext), extra={
        #     'AppName': 'Campaign'})
        # return check_malicious_file_from_filename(file_ext, allowed_files_list)
    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("In check_malicious_file_from_content: %s at %s", str(e), str(exc_tb.tb_lineno), extra={
            'AppName': 'Campaign'})
        return True


def validate_mobile_number(mobile_number):
    try:
        reg = r'[6-9][0-9]{9}'
        if re.fullmatch(reg, mobile_number):
            return True

        return False
    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("In validate_mobile_number: %s at %s", str(e), str(exc_tb.tb_lineno), extra={
            'AppName': 'Campaign'})

    return False


def check_url_valid(url):
    try:
        form_obj = forms.URLField()
        form_obj.clean(url)
        return True
    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("In check_url_valid: %s at %s", str(e), str(exc_tb.tb_lineno), extra={
            'AppName': 'Campaign'})

    return False


def get_campaign_last_saved_state(campaign_obj):
    last_saved_state = None
    try:
        if campaign_obj.last_saved_state is not None:
            last_saved_state = campaign_obj.last_saved_state
    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("In get_campaign_last_saved_state: %s at %s", str(e), str(exc_tb.tb_lineno), extra={
            'AppName': 'Campaign'})

    return last_saved_state


def check_campaign_current_page_valid(current_page_state, campaign_obj):
    data = {
        "is_valid": False,
        "last_saved_state": None,
    }

    try:
        last_saved_state = get_campaign_last_saved_state(campaign_obj)
        data["last_saved_state"] = last_saved_state

        if last_saved_state == None:
            data["is_valid"] = False
        else:
            if PAGE_ORDER[current_page_state] - PAGE_ORDER[last_saved_state] > 1:
                data['is_valid'] = False
            else:
                data['is_valid'] = True

    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("In check_campaign_current_page_valid: %s at %s", str(e), str(exc_tb.tb_lineno), extra={
            'AppName': 'Campaign'})

    return data


def get_last_saved_page_url(last_saved_state, bot_pk, campaign_id):
    try:
        if last_saved_state == None:
            url = '/campaign/create-campaign/?bot_pk=' + str(bot_pk)
        elif last_saved_state == CAMPAIGN_BASIC_INFO_STATE:
            url = '/campaign/create-campaign/?bot_pk=' + str(bot_pk)
        elif last_saved_state == CAMPAIGN_TAG_AUDIENCE_STATE:
            url = '/campaign/tag-audience/?bot_pk=' + str(bot_pk)
        elif last_saved_state == CAMPAIGN_TEMPLATE_STATE:
            url = '/campaign/create-template/?bot_pk=' + str(bot_pk)
        elif last_saved_state == CAMPAIGN_SETTINGS_STATE:
            url = '/campaign/voice-bot/settings/?bot_pk=' + str(bot_pk)
        elif last_saved_state == CAMPIGN_REVIEW_STATE:
            url = '/campaign/review/?bot_pk=' + str(bot_pk)
        elif last_saved_state == CAMPAIGN_VB_REVIEW_STATE:
            url = '/campaign/voice-bot/review/?bot_pk=' + str(bot_pk)
        else:
            # Show create campaign page in case of unknow saved state
            url = '/campaign/create-campaign/?bot_pk=' + str(bot_pk)

        if campaign_id is not None:
            url += "&campaign_id=" + str(campaign_id)

        return url
    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("In get_last_saved_page_url: %s at %s", str(e), str(exc_tb.tb_lineno), extra={
            'AppName': 'Campaign'})
        return None


def check_campaign_in_progress(campaign_obj):
    try:
        if campaign_obj.status == CAMPAIGN_IN_PROGRESS:
            return True
        return False
    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("In check_campaign_in_progress: %s at %s", str(e), str(exc_tb.tb_lineno), extra={
            'AppName': 'Campaign'})
        return True


def sample_data_wrapper_function(file_path, Profile, campaign_channel, sheet, original_file_name, event_obj, bot_obj, status_message, metadata):
    try:
        try:
            func_timeout.func_timeout(
                CAMPAIGN_MAXIMUM_SHEET_UPLOAD_TIME_LIMIT, get_sample_data_from_batch_file, args=[file_path, Profile, campaign_channel, sheet, original_file_name, event_obj, bot_obj, status_message, metadata])
        except func_timeout.FunctionTimedOut:
            os.remove(file_path)
            status = 400
            row_errors = [
                "The process took longer than expected. Please try again later."]
            total_batch_count = metadata['max_row'] - 1
            event_obj.event_info = json.dumps({"row_errors": row_errors,
                                               "status": status,
                                               "original_file_name": original_file_name,
                                               "header_metadata": [],
                                               "sample_data": [],
                                               "total_batch_count": total_batch_count,
                                               "total_opted_in": 0,
                                               "status_message": status_message
                                               })
            event_obj.is_completed = True
            event_obj.event_progress = 0
            event_obj.save()

    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("In sample_data_wrapper_function: %s at %s", str(e), str(exc_tb.tb_lineno), extra={
            'AppName': 'Campaign'})


def get_sample_data_from_batch_file(file_path, Profile, campaign_channel, sheet, original_file_name, event_obj, bot_obj, status_message, metadata):
    try:
        row_errors, header_metadata, sample_data, status = [], [], [], None
        max_row, max_column = metadata['max_row'], metadata['max_column']
        auto_delete_checked = metadata['auto_delete_checked']
        error_rows = metadata['error_rows']
        error_found, total_opted_in, deleted_rows = False, 0, 0

        first_col_name = sheet.cell_value(0, 0)
        header_data = []
        for col in range(max_column):
            header_name = sheet.cell_value(0, col)
            if not header_name:
                header_name = ''
            header_name = remo_html_from_string(header_name)
            header_metadata.append({
                'col_no': col,
                'col_name': header_name
            })
            header_data.append(header_name)

        if 'phone' not in first_col_name.lower():
            status_message = 'First column should always be for the Phone.  Please change the column name ' + first_col_name + \
                ' to Phone or you can download and use our Template Batch file to upload the contacts in the prescribed text format only'
            status = 402
        elif '' in header_data:
            if (header_data.count('') > 1):
                status_message = "Two or more header names are empty or have no values filled in the columns of the uploaded file. Please enter the header names or remove the columns if not required and upload the file again."
            else:
                status_message = f"Empty header name detected in a column number {str(header_data.index('') + 1)} of the uploaded file. Please enter the header name or remove the column if not required and upload the file again."
            status = 402
        elif len(set(header_data)) != len(header_data):
            duplicate_column = list(
                set([x for x in header_data if header_data.count(x) > 1]))
            if (len(duplicate_column) > 1):
                status_message = f"We have detected the same header names {' and '.join(duplicate_column)} in the columns of the uploaded file. Please change the repeated header names with the unique names and upload the file again."
            else:
                status_message = f"We have detected the same header name {duplicate_column[0]} in the column of the uploaded file. Please change the repeated header name with a unique name and upload the file again."
            status = 402
        else:
            if auto_delete_checked:
                deletion_metadata = {"error_rows": error_rows, "header_data": header_data,
                                     "max_row": max_row, "max_col": max_column}

                max_row, sheet, valid_workbook, total_opted_in, sample_data = delete_invalid_rows_of_sheet(
                    sheet, deletion_metadata, event_obj, Profile)

                valid_workbook.save(file_path)
                deleted_rows = len(error_rows)
                error_rows = []

            else:
                error_workbook = openpyxl.Workbook(write_only=True)
                error_sheet = error_workbook.create_sheet()
                error_sheet.append(header_data)
                error_data = []

                for row in range(1, max_row):
                    row_data = [str(sheet.cell_value(row, col))
                                for col in range(max_column)]
                    if not ''.join(row_data):
                        continue
                    
                    progress = int((row - 1) / (max_row - 1) * 100)
                    if row % 500 == 0:
                        event_obj.event_progress = progress
                        event_obj.save(update_fields=["event_progress"])

                    mobile_number = sheet.cell_value(row, 0)
                    mobile_number, error_found = identify_invalid_phone_numbers(
                        mobile_number, campaign_channel, row)
                    if error_found:
                        error_rows.append(row)
                        row_errors.append(error_found)
                        for col in range(max_column):
                            error_data.append(sheet.cell_value(row, col))
                        error_data.append(error_found)

                    profile_obj = Profile.objects.filter(user_id=mobile_number)
                    if profile_obj and profile_obj[0].campaign_optin:
                        total_opted_in += 1

                    # To get sample batch data to show on modal
                    if not error_rows and row < 4:
                        sample_data.append(
                            get_sample_data(sheet, max_column, row))

                    if error_found:
                        error_sheet.append(error_data)
                        error_data = []

                if error_rows:
                    error_workbook.save(file_path)
                else:
                    error_sheet.close()

        if max_row < 2:
            status = 403

        total_batch_count = max_row - 1
        response = {}
        file_access_management_obj = CampaignFileAccessManagement.objects.create(
            file_path=file_path, is_public=False, original_file_name=original_file_name, bot=bot_obj)
        response["file_path"] = "/campaign/download-file/" + \
            str(file_access_management_obj.key) + "/"

        if row_errors or status in [402, 403]:
            if not status:
                status = 400
            row_errors = row_errors[:CAMPAIGN_MAXIMUM_SCROLLER_LIST_LIMIT]
            event_obj.event_info = json.dumps({"row_errors": row_errors,
                                               "status": status,
                                               "original_file_name": original_file_name,
                                               "header_metadata": header_metadata,
                                               "sample_data": sample_data,
                                               "total_batch_count": total_batch_count,
                                               "total_opted_in": total_opted_in,
                                               "file_path": response["file_path"],
                                               "error_rows": error_rows,
                                               "status_message": status_message
                                               })
        else:
            status = 200
            message = "success"
            if campaign_channel == "RCS":
                rcs_enabled_users = 0
                rcs_enabled_users, status, status_message = add_rcs_enabled_column(
                    file_path, bot_obj, RCSDetails)
                if status == 200:
                    total_opted_in = rcs_enabled_users

            event_obj.event_info = json.dumps({"row_errors": row_errors[:CAMPAIGN_MAXIMUM_SCROLLER_LIST_LIMIT],
                                               "status": status,
                                               "original_file_name": original_file_name,
                                               "header_metadata": header_metadata,
                                               "sample_data": sample_data,
                                               "total_batch_count": total_batch_count,
                                               "total_opted_in": total_opted_in,
                                               "file_path": response["file_path"],
                                               "error_rows": error_rows,
                                               "status_message": status_message,
                                               "deleted_rows": deleted_rows,
                                               "message": message
                                               })
        event_obj.is_completed = True
        event_obj.event_progress = 100
        event_obj.completed_datetime = datetime.now()
        event_obj.save()

    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("In get_sample_data_from_batch_file: %s at %s", str(e), str(exc_tb.tb_lineno), extra={
            'AppName': 'Campaign'})

        event_obj.is_completed = True
        event_obj.completed_datetime = datetime.now()
        event_obj.save()


def identify_invalid_phone_numbers(mobile_number, campaign_channel, row):
    try:
        error_found = ""
        row += 1
        if not mobile_number:
            error_found = 'Row ' + \
                str(row) + ' : Phone Number field cannot be empty, please add a valid Phone number with the country code eg: 9194194XXXX and Re-upload the file'
            return mobile_number, error_found

        mobile_number = remo_html_from_string(
            str(mobile_number))

        if campaign_channel != "RCS" and campaign_channel != "Whatsapp Business":
            mobile_number = remo_special_tag_from_string(
                mobile_number)

        mobile_number = mobile_number.strip()
        if mobile_number == "":
            error_found = 'Row ' + \
                str(row) + ' : Phone Number field cannot be empty, please add a valid Phone number with the country code eg: 9194194XXXX and Re-upload the file'
        else:
            if mobile_number[-2:] == ".0":
                mobile_number = mobile_number.replace('.0', '')
            mobile_number = validation_obj.removing_phone_non_digit_element(
                mobile_number)
            if campaign_channel == "RCS" or campaign_channel == "Whatsapp Business":
                if mobile_number[0] != "+":
                    validation_mobile_number = "+" + mobile_number
                else:
                    validation_mobile_number = mobile_number
                if not re.match(INTERNATION_PHONE_NUMBER_REGEX, validation_mobile_number):
                    error_found = 'Row ' + \
                        str(row) + ' : Invalid Phone Number. The phone number present is ' + mobile_number + \
                        '. Please check if the phone number has a valid format and country code eg: 9194194XXXX. You can download our Batch File and use that in the prescribed text format only.'
            else:
                if not validate_mobile_number(mobile_number):
                    error_found = 'Row ' + str(row) + ' : Invalid Phone Number. The phone number present is ' + mobile_number + \
                        '. Please check if the phone number has a valid format and without country code eg: 94194XXXXX. You can download our Batch File and use that in the prescribed text format only.'

    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("In identify_invalid_phone_numbers: %s at %s", str(e), str(exc_tb.tb_lineno), extra={
            'AppName': 'Campaign'})
    return mobile_number, error_found


def get_sample_data(sheet, max_column, row):
    try:
        data = []
        for col in range(max_column):
            value = sheet.cell_value(row, col)
            if not value:
                value = "-"
            else:
                value = clean_input(str(value))
                if col == 0:
                    if value[-2:] == ".0":
                        value = value.replace('.0', '')
                    value = int(
                        validation_obj.removing_phone_non_digit_element(value))
            data.append(value)

    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("In get_sample_data: %s at %s", str(e), str(exc_tb.tb_lineno), extra={
            'AppName': 'Campaign'})

    return data


def delete_invalid_rows_of_sheet(sheet, deletion_metadata, event_obj, Profile):
    try:
        valid_workbook = openpyxl.Workbook(write_only=True)
        new_sheet = valid_workbook.create_sheet()
        new_sheet.title = sheet.name
        changed_rows, total_opted_in, sample_data = 0, 0, []
        max_row, max_col = deletion_metadata["max_row"], deletion_metadata["max_col"]
        error_rows, header_data = deletion_metadata["error_rows"], deletion_metadata["header_data"]

        # Adding headers
        new_sheet.append(header_data)

        for row in range(1, max_row):
            progress = int((row - 1) / (max_row - 1) * 100)
            if row % 500 == 0:
                event_obj.event_progress = progress
                event_obj.save(update_fields=["event_progress"])
            if row in error_rows:
                changed_rows += 1
                continue
            valid_row = []
            for col in range(0, max_col):
                valid_row.append(sheet.cell_value(row, col))

            new_sheet.append(valid_row)

            mobile_number = sheet.cell_value(row - changed_rows, 0)
            profile_obj = Profile.objects.filter(user_id=mobile_number)
            if profile_obj and profile_obj[0].campaign_optin:
                total_opted_in += 1

            if 0 < row - changed_rows < 4:
                sample_data.append(get_sample_data(
                    sheet, max_col, row))

    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("In delete_invalid_rows_of_sheet: %s at %s", str(e), str(exc_tb.tb_lineno), extra={
            'AppName': 'Campaign'})

    return max_row - changed_rows, new_sheet, valid_workbook, total_opted_in, sample_data


def delete_file_by_key(key, CampaignFileAccessManagement):
    try:
        file_access_management_obj = CampaignFileAccessManagement.objects.filter(
            key=key)

        if not file_access_management_obj:
            return False
        else:
            file_path = file_access_management_obj[0].file_path
            os.remove(file_path)
            return True

    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("In delete_file_by_key: %s at %s", str(e), str(exc_tb.tb_lineno), extra={
            'AppName': 'Campaign'})

        return False


def get_active_campaign_usign_batch_id(batch_obj):
    try:
        campaigns = batch_obj.campaigns.all()

        for campaign in campaigns:
            if campaign.status == CAMPAIGN_IN_PROGRESS:
                return True, campaign.name
        campaign_schedule = CampaignScheduleObject.objects.filter(campaign_batch=batch_obj).first()
        if campaign_schedule:
            return True, campaign_schedule.campaign.name
        return False, ''
    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("In get_active_campaign_usign_batch_id: %s at %s", str(e), str(exc_tb.tb_lineno), extra={
            'AppName': 'Campaign'})

        return False, ''


def get_active_campaign_usign_template(template_obj, campaign_channel, Campaign):
    try:
        if campaign_channel == 'RCS':
            campaigns = Campaign.objects.filter(
                campaign_template_rcs=template_obj)
        else:
            campaigns = Campaign.objects.filter(
                campaign_template=template_obj)

        for campaign in campaigns:
            if campaign.status == CAMPAIGN_IN_PROGRESS:
                return (True, campaign.name)

        return (False, '')
    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("In get_active_campaign_usign_template: %s at %s", str(e), str(exc_tb.tb_lineno), extra={
            'AppName': 'Campaign'})

        return False, ''


def template_input_validations(template_data, row):
    try:
        row_errors = []
        error_found = False
        
        template_type = template_data['template_type']
        template_name = template_data['template_name']
        template_category = template_data['template_category']
        template_body = template_data['template_body']
        template_header = template_data['template_header']
        template_cta_text = template_data['template_cta_text']
        template_cta_link = template_data['template_cta_link']
        template_attachment = template_data['template_attachment'].strip()
        template_button_type = template_data['template_button_type']
        template_callus_text = template_data['template_callus_text']
        template_callus_number = template_data['template_callus_number']
        template_qr_1 = template_data['template_qr_1']
        template_qr_2 = template_data['template_qr_2']
        template_qr_3 = template_data['template_qr_3']
        type_of_first_cta_btton = template_data['type_of_first_cta_btton']
        document_file_name = template_data['document_file_name']

        # checking if user have selected from dropdown from template
        if template_type not in ['text', 'video', 'document', 'image']:
            error_found = True
            error_message = "Row %s: Selected 'Type' should selected from dropdown('text', 'video', 'document', 'image'), please check the template." % str(
                row + 1)
            row_errors.append(error_message)

        if template_button_type and template_button_type not in ['none', 'cta', 'quick_reply']:
            error_found = True
            error_message = "Row %s: Selected 'Button Type' should selected from dropdown('none', 'cta', 'quick_reply'), please check the template." % str(
                row + 1)
            row_errors.append(error_message)
        if type_of_first_cta_btton and type_of_first_cta_btton not in ['none', 'website_link', 'call_us']:
            error_found = True
            error_message = "Row %s: Selected 'CTA 1st Button Type' should selected from dropdown('none', 'website_link', 'call_us'), please check the template." % str(
                row + 1)
            row_errors.append(error_message)

        # Attachment validation on template type
        if template_type == 'text':
            if template_attachment != '':
                error_found = True
                error_message = "Row %s: Selected message type 'Text' does not expect an URL in the Attachment URL column, please check the template." % str(
                    row + 1)
                row_errors.append(error_message)
            if template_header != clean_input(template_header):
                error_found = True
                error_message = "Row %s: There are some vulnerability in Message header template, please check the template." % str(
                    row + 1)
                row_errors.append(error_message)
            hearder_variable = re.findall(r'\{\{(.*?)\}\}', template_header)
            if hearder_variable:
                hearder_variable = hearder_variable[0]
            else:
                hearder_variable = ''
            if len(template_header.replace('{{' + hearder_variable + '}}', '_')) > 60:
                error_found = True
                error_message = "Row %s: Message header template length cannot be greater than 60." % str(
                    row + 1)
                row_errors.append(error_message)
        elif template_type in ['video', 'image', 'document']:
            if template_attachment == '':
                error_found = True
                error_message = f"Row %s: Selected message type '{template_type.capitalize()}' expects an URL in the Attachment URL column, please check the template." % str(
                    row + 1)
                row_errors.append(error_message)
            if not check_url_valid(template_attachment):
                error_found = True
                error_message = "Row %s: Please fill valid URL in Attachment URL column, please check the template." % str(
                    row + 1)
                row_errors.append(error_message)
            if template_header:
                error_found = True
                error_message = f"Row %s: Selected message type '{template_type.capitalize()}' does not expect any Text message in the 'Message Header Template' column, please check the template." % str(
                    row + 1)
                row_errors.append(error_message)
            if template_type == "document":
                if len(document_file_name) > 100:
                    error_found = True
                    error_message = "Row %s: The 'Document File Name' can't be greater than 100 characters." % str(
                        row + 1)
                    row_errors.append(error_message)
                if document_file_name.strip() and not validation_obj.is_filename_alphanumeric(document_file_name):
                    error_found = True
                    error_message = "Row %s: Make sure that the 'Document File Name' column must have a valid name, The name format should have atleast one alphanumeric character and if required the supported special characters are ['(', ')', '@', '-', '_', '=']." % str(
                        row + 1)
                    row_errors.append(error_message)

        # Button type validation on template
        if template_button_type == 'none':
            if template_callus_number != "" or template_cta_link != "":
                error_found = True
                error_message = "Row %s: Selected button type 'None' does not expect any values in Website URL or Phone Number, please check the template." % str(
                    row + 1)
                row_errors.append(error_message)
            elif template_qr_1 != "" or template_qr_2 != "" or template_qr_3 != "":
                error_found = True
                error_message = "Row %s: Selected button type 'None' does not expect any values in Quick Reply columns, please check the template." % str(
                    row + 1)
                row_errors.append(error_message)
        elif template_button_type == 'cta':  # Validation for button type is selected cta
            if template_callus_text != "" and len(template_callus_text) > 25:
                error_found = True
                error_message = "Row %s: Call Us Button Text length cannot be greater than 25." % str(
                    row + 1)
                row_errors.append(error_message)
            if template_cta_text != "" and len(template_cta_text) > 25:
                error_found = True
                error_message = "Row %s: Website Button Text length cannot be greater than 25." % str(
                    row + 1)
                row_errors.append(error_message)
            template_cta_link_variable = re.findall(r'\{\{(.*?)\}\}', template_cta_link)
            if template_cta_link_variable:
                template_cta_link_variable = '{{' + template_cta_link_variable[0] + '}}'
            if template_cta_link != "":
                if (template_cta_link_variable and template_cta_link_variable in template_cta_link and len(template_cta_link) - len(template_cta_link_variable) > 2000) or len(template_cta_link) > 2000:
                    error_found = True
                    error_message = "Row %s: Websit URL length cannot be greater than 2000." % str(
                        row + 1)
                    row_errors.append(error_message)
            if template_callus_number != "" and len(template_callus_number) > 20:
                error_found = True
                error_message = "Row %s: Phone Number length cannot be greater than 20." % str(
                    row + 1)
                row_errors.append(error_message)
            if template_cta_link != '' and template_cta_link_variable and template_cta_link_variable in template_cta_link:
                if template_cta_link[-len(template_cta_link_variable):] != template_cta_link_variable:
                    error_found = True
                    error_message = "Row %s: The URL to be added is used for the CTA button, in the case of Dynamic URL we need to pass the full URL with the variable at the end. Eg: https://www.youtube.com/{{1}}." % str(
                        row + 1)
                    row_errors.append(error_message)
                if not check_url_valid(template_cta_link[:-len(template_cta_link_variable)]):
                    error_found = True
                    error_message = "Row %s: Please fill valid Website URL in Website URL columns, please check the template." % str(
                        row + 1)
                    row_errors.append(error_message)
            elif template_cta_link != '':
                if not check_url_valid(template_cta_link):
                    error_found = True
                    error_message = "Row %s: Please fill valid Website URL in Website URL columns, please check the template." % str(
                        row + 1)
                    row_errors.append(error_message)
            if template_callus_number == '' and template_cta_link == '':
                error_found = True
                error_message = "Row %s: Selected button type 'CTA' expects values in Website URL or Phone Number columns, please check the template." % str(
                    row + 1)
                row_errors.append(error_message)
            if (template_callus_number != '' and template_callus_text == '') or (template_callus_number == '' and template_callus_text != ''):
                error_found = True
                error_message = "Row %s: Please fill both Call Us Button Text and Phone Number columns, please check the template." % str(
                    row + 1)
                row_errors.append(error_message)
            if (template_cta_link != '' and template_cta_text == '') or (template_cta_link == '' and template_cta_text != ''):
                error_found = True
                error_message = "Row %s: Please fill both Website Button Text and Website URL columns, please check the template." % str(
                    row + 1)
                row_errors.append(error_message)
            if template_callus_number:
                template_callus_number = validation_obj.removing_phone_non_digit_element(template_callus_number)
                if '+' not in template_callus_number:
                    template_callus_number = '+' + template_callus_number
                if not re.match(INTERNATION_PHONE_NUMBER_REGEX, template_callus_number):
                    error_found = True
                    error_message = "Row %s: Invalid Mobile Number (Please check if phone number has a valid format and country code eg: +9194194XXXX)" % str(
                        row + 1)
                    row_errors.append(error_message)
        elif template_button_type == 'quick_reply':  # Validation for button type is selected quick_reply
            if template_qr_1 != '' and len(template_qr_1) > 25:
                error_found = True
                error_message = "Row %s: Quick Reply 1 cannot be greater than 25." % str(
                    row + 1)
                row_errors.append(error_message)
            if template_qr_2 != '' and len(template_qr_2) > 25:
                error_found = True
                error_message = "Row %s: Quick Reply 2 cannot be greater than 25." % str(
                    row + 1)
                row_errors.append(error_message)
            if template_qr_3 != '' and len(template_qr_3) > 25:
                error_found = True
                error_message = "Row %s: Quick Reply 3 cannot be greater than 25." % str(
                    row + 1)
                row_errors.append(error_message)
            if template_qr_1 == '' and template_qr_2 == '' and template_qr_3 == '':
                error_found = True
                error_message = "Row %s: Selected button type 'Quick Reply' expects Text values in the Quicky Reply columns, please check the template." % str(
                    row + 1)
                row_errors.append(error_message)

        if template_type == "":
            error_found = True
            error_message = "Row %s: Template type is not empty" % str(
                row + 1)
            row_errors.append(error_message)

        if template_name == "":
            error_found = True
            error_message = "Row %s: Template name is empty" % str(row + 1)
            row_errors.append(error_message)

        if template_category == "":
            error_found = True
            error_message = "Row %s: Template category is empty" % str(
                row + 1)
            row_errors.append(error_message)

        if template_body == "":
            error_found = True
            error_message = "Row %s: Template body is empty" % str(row + 1)
            row_errors.append(error_message)
        return error_found, row_errors
    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("In template_input_validations: %s at %s", str(e), str(exc_tb.tb_lineno), extra={
            'AppName': 'Campaign'})
        return True, ['Something Went Wrong']


def get_sample_data_from_template_file(file_path, bot_obj):
    try:
        ensure_element_tree(xlrd)
        status_code = 400
        wb = xlrd.open_workbook(file_path)
        sheet = wb.sheet_by_index(0)

        row_errors = []
        sample_data = []
        template_exist_rows = []

        column_header_list = [
            "type",
            "language",
            "template name",
            "category",
            "message body template",
            "message header template",
            "message footer template",
            "button type",
            "cta 1st button type",
            "website button text",
            "website url",
            "call us button text",
            "phone number",
            "quick reply 1",
            "quick reply 2",
            "quick reply 3",
            "attachment url",
            "document file name"
        ]

        is_column_mismatched = False

        for col in range(0, sheet.ncols):
            header_name = sheet.cell_value(0, col)
            header_name = header_name.replace('*', '')
            header_name = header_name.lower()
            if col >= len(column_header_list) or header_name != column_header_list[col]:
                is_column_mismatched = True
                break

        if is_column_mismatched:
            error_message = "Uploaded file column names does not match with the expected file, please download and upload the latest template file."
            row_errors.append(error_message)
            status_code = 400
            return status_code, row_errors, sample_data, []

        is_file_empty = True
        for row in range(1, sheet.nrows):
            error_found = False
            is_file_empty = False

            template_type = sheet.cell_value(row, 0)
            template_language = sheet.cell_value(row, 1)
            template_name = sheet.cell_value(row, 2)
            template_category = sheet.cell_value(row, 3)
            template_body = sheet.cell_value(row, 4)
            template_header = sheet.cell_value(row, 5)
            template_footer = sheet.cell_value(row, 6)
            template_button_type = sheet.cell_value(row, 7)
            type_of_first_cta_btton = sheet.cell_value(row, 8)
            template_cta_text = sheet.cell_value(row, 9)
            template_cta_link = sheet.cell_value(row, 10)
            template_callus_text = sheet.cell_value(row, 11)
            template_callus_number = sheet.cell_value(row, 12)
            template_qr_1 = sheet.cell_value(row, 13)
            template_qr_2 = sheet.cell_value(row, 14)
            template_qr_3 = sheet.cell_value(row, 15)
            template_attachment = sheet.cell_value(row, 16)
            try:
                document_file_name = sheet.cell_value(row, 17)
            except:
                document_file_name = CAMPAIGN_DOCUMENT_FILE_NAME

            template_name = str(template_name).strip()
            template_language = str(template_language).lower().strip()
            template_category = str(template_category).lower().strip()
            template_header = str(template_header).strip()
            template_footer = str(template_footer).strip()
            template_type = str(template_type).lower().strip()
            template_cta_text = str(template_cta_text).strip()
            template_cta_link = str(template_cta_link).strip()
            template_button_type = str(template_button_type).lower().strip()
            template_callus_text = str(template_callus_text).strip()
            template_qr_1 = str(template_qr_1).strip()
            template_qr_2 = str(template_qr_2).strip()
            template_qr_3 = str(template_qr_3).strip()
            template_attachment = str(template_attachment).strip()
            type_of_first_cta_btton = str(type_of_first_cta_btton).lower().strip()
            template_callus_number = str(template_callus_number)
            if template_callus_number[-2:] == ".0":
                template_callus_number = template_callus_number.replace('.0', '')
            document_file_name = str(document_file_name).strip()

            template_type = clean_input(template_type)
            template_language = clean_input(template_language)
            template_name = clean_input(template_name)
            template_category = clean_input(template_category)
            template_cta_text = clean_input(template_cta_text)
            template_cta_link = clean_input(template_cta_link)
            template_attachment = clean_input(template_attachment)
            template_button_type = clean_input(template_button_type)
            template_callus_text = clean_input(template_callus_text)
            template_qr_1 = clean_input(template_qr_1)
            template_qr_2 = clean_input(template_qr_2)
            template_qr_3 = clean_input(template_qr_3)
            type_of_first_cta_btton = clean_input(type_of_first_cta_btton)

            template_data = {
                'template_type': template_type,
                'template_name': template_name,
                'template_category': template_category,
                'template_body': template_body,
                'template_header': template_header,
                'template_cta_text': template_cta_text,
                'template_cta_link': template_cta_link,
                'template_attachment': template_attachment,
                'template_button_type': template_button_type,
                'template_callus_text': template_callus_text,
                'template_callus_number': template_callus_number,
                'template_qr_1': template_qr_1,
                'template_qr_2': template_qr_2,
                'template_qr_3': template_qr_3,
                'type_of_first_cta_btton': type_of_first_cta_btton,
                'document_file_name': document_file_name,
            }

            error_found, row_errors = template_input_validations(template_data, row)
            template_cta_link_variable = re.findall(r'\{\{(.*?)\}\}', template_cta_link)
            original_template_cta_link = template_cta_link
            if template_cta_link_variable:
                template_cta_link_variable = '{{' + template_cta_link_variable[0] + '}}'
                template_cta_link = template_cta_link[:-len(template_cta_link_variable)]
            if not check_url_valid(template_cta_link):
                template_cta_link = None
            else:
                template_cta_link = original_template_cta_link

            if not check_url_valid(template_attachment):
                template_attachment = None

            if not error_found and row < 3:
                status_code = 200
                data = {
                    "type": template_type,
                    "language": template_language,
                    "template_name": template_name,
                    "category": template_category,
                    "template_body": template_body,
                    "template_header": template_header,
                    "template_footer": template_footer,
                    "template_cta_text": template_cta_text,
                    "template_cta_link": template_cta_link,
                    "template_attachment": template_attachment,
                    "template_button_type": template_button_type,
                    "template_callus_text": template_callus_text,
                    "template_callus_number": template_callus_number,
                    "template_qr_1": template_qr_1,
                    "template_qr_2": template_qr_2,
                    "template_qr_3": template_qr_3,
                    "document_file_name": document_file_name,
                }
                sample_data.append(data)
            else:
                status_code = 400
            template_name = str(template_name).lower().strip()
            template_language = str(template_language).lower().strip()
            
            language_obj = CampaignTemplateLanguage.objects.filter(
                title=template_language).first()
            if language_obj is None:
                language_obj = CampaignTemplateLanguage.objects.create(
                    title=template_language)
            
            campaign_template_obj = CampaignTemplate.objects.filter(
                template_name=template_name, language=language_obj, is_deleted=False, bot=bot_obj).first()
            # if template which have same name and same language will not get created
            if campaign_template_obj:
                template_exist_rows.append(row)

        if is_file_empty:
            status_code = 403
            row_errors.append('File has no data to upload')
        return status_code, row_errors, sample_data, template_exist_rows
    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("In get_sample_data_from_template_file: %s at %s", str(e), str(exc_tb.tb_lineno), extra={
            'AppName': 'Campaign'})

        return status_code, ['Some errors occurred while reading the template file. Please make sure all the column headers are matching with the Sample Template.'], [], []


def create_campaign_templates_from_file(file_path, file_name, file_key, bot_obj, CampaignTemplate, CampaignTemplateLanguage, CampaignTemplateCategory, CampaignTemplateStatus, CampaignTemplateType):
    error_data = {
        "template_name_exist": [],
        "error": None,
        # "template_id_exist": [],
    }
    try:
        ensure_element_tree(xlrd)

        wb = xlrd.open_workbook(file_path)
        sheet = wb.sheet_by_index(0)

        template_status_obj = CampaignTemplateStatus.objects.filter(
            title="approved").first()

        template_name_exist_list = []
        template_language_exist_list = []
        template_exist_rows = []
        saving_file_path = '/campaign/download-file/' + file_key + '/' + file_name
        for row in range(1, sheet.nrows):
            template_type = sheet.cell_value(row, 0)
            template_language = sheet.cell_value(row, 1)
            template_name = sheet.cell_value(row, 2)
            template_category = sheet.cell_value(row, 3)
            template_body = sheet.cell_value(row, 4)
            template_header = sheet.cell_value(row, 5)
            template_footer = sheet.cell_value(row, 6)
            template_button_type = sheet.cell_value(row, 7)
            type_of_first_cta_btton = sheet.cell_value(row, 8)
            template_cta_text = sheet.cell_value(row, 9)
            template_cta_link = sheet.cell_value(row, 10)
            template_callus_text = sheet.cell_value(row, 11)
            template_callus_number = sheet.cell_value(row, 12)
            template_qr_1 = sheet.cell_value(row, 13)
            template_qr_2 = sheet.cell_value(row, 14)
            template_qr_3 = sheet.cell_value(row, 15)
            template_attachment = sheet.cell_value(row, 16)
            try:
                document_file_name = sheet.cell_value(row, 17)
            except:
                document_file_name = CAMPAIGN_DOCUMENT_FILE_NAME

            template_name = str(template_name).lower().strip()
            template_language = str(template_language).lower().strip()
            template_category = str(template_category).lower().strip()
            template_type = str(template_type).lower().strip()
            template_cta_text = str(template_cta_text).strip()
            template_button_type = str(template_button_type).lower().strip()
            template_callus_text = str(template_callus_text).strip()
            template_qr_1 = str(template_qr_1).strip()
            template_qr_2 = str(template_qr_2).strip()
            template_qr_3 = str(template_qr_3).strip()
            type_of_first_cta_btton = str(type_of_first_cta_btton).lower().strip()
            template_callus_number = str(template_callus_number)
            if template_callus_number[-2:] == ".0":
                template_callus_number = template_callus_number.replace('.0', '')
            document_file_name = str(document_file_name).strip()

            template_type = remo_html_from_string(template_type)
            template_language = remo_html_from_string(template_language)
            template_name = remo_html_from_string(template_name)
            template_category = remo_html_from_string(template_category)
            template_body = validation_obj.sanitize_html(template_body)
            template_header = validation_obj.sanitize_html(template_header)
            template_footer = validation_obj.sanitize_html(template_footer)
            template_cta_text = remo_html_from_string(template_cta_text)
            template_cta_link = remo_html_from_string(template_cta_link)
            template_attachment = remo_html_from_string(template_attachment).strip()
            template_button_type = remo_html_from_string(template_button_type)
            template_callus_text = remo_html_from_string(template_callus_text)
            template_qr_1 = remo_html_from_string(template_qr_1)
            template_qr_2 = remo_html_from_string(template_qr_2)
            template_qr_3 = remo_html_from_string(template_qr_3)
            template_callus_number = remo_html_from_string(template_callus_number)
            type_of_first_cta_btton = remo_html_from_string(type_of_first_cta_btton)

            template_type = remo_special_tag_from_string(template_type)
            template_language = remo_special_tag_from_string(template_language)
            template_name = remo_special_tag_from_string(template_name)
            template_category = remo_special_tag_from_string(template_category)
            type_of_first_cta_btton = remo_special_tag_from_string(type_of_first_cta_btton)
            
            language_obj = CampaignTemplateLanguage.objects.filter(
                title=template_language).first()
            if language_obj is None:
                language_obj = CampaignTemplateLanguage.objects.create(
                    title=template_language)
            
            campaign_template_obj = CampaignTemplate.objects.filter(
                template_name=template_name, language=language_obj, is_deleted=False, bot=bot_obj).first()
            # if template which have same name and same language will not get created
            if campaign_template_obj:
                template_name_exist_list.append(template_name)
                template_language_exist_list.append(language_obj.title)
                template_exist_rows.append(row)
                continue

            if not check_url_valid(template_attachment):
                template_attachment = None

            template_cta_link_variable = re.findall(r'\{\{(.*?)\}\}', template_cta_link)
            original_template_cta_link = template_cta_link
            if template_cta_link_variable:
                template_cta_link_variable = '{{' + template_cta_link_variable[0] + '}}'
                template_cta_link = template_cta_link[:-len(template_cta_link_variable)]
            if not check_url_valid(template_cta_link):
                template_cta_link = None
            else:
                template_cta_link = original_template_cta_link

            if template_button_type == 'cta':
                template_qr_1 = None
                template_qr_2 = None
                template_qr_3 = None
            elif template_button_type == 'quick_reply':
                template_callus_text = None
                template_callus_number = None
                template_cta_text = None
                template_cta_link = None
                type_of_first_cta_btton = None
            else:
                template_callus_text = None
                template_callus_number = None
                template_cta_text = None
                template_cta_link = None
                template_qr_1 = None
                template_qr_2 = None
                template_qr_3 = None
                type_of_first_cta_btton = None
            template_metadata = {}
            template_metadata['button_type'] = template_button_type
            template_metadata['callus_text'] = template_callus_text
            template_metadata['callus_number'] = template_callus_number
            template_metadata['template_qr_1'] = template_qr_1
            template_metadata['template_qr_2'] = template_qr_2
            template_metadata['template_qr_3'] = template_qr_3
            template_metadata['type_of_first_cta_btton'] = type_of_first_cta_btton
            template_metadata['template_file_path'] = saving_file_path
            template_metadata['document_file_name'] = document_file_name
            template_metadata = json.dumps(template_metadata)
            campaign_template_obj = CampaignTemplate.objects.create(
                template_name=template_name,
                template_header=template_header,
                template_body=template_body,
                template_footer=template_footer,
                cta_text=template_cta_text,
                cta_link=template_cta_link,
                attachment_src=template_attachment,
                template_metadata=template_metadata)

            category_obj = CampaignTemplateCategory.objects.filter(
                title=template_category).first()
            if category_obj is None:
                category_obj = CampaignTemplateCategory.objects.create(
                    title=template_category)

            template_type_obj = CampaignTemplateType.objects.filter(
                title=template_type).first()
            if template_type_obj is None:
                template_type_obj = CampaignTemplateType.objects.create(
                    title=template_type)

            campaign_template_obj.language = language_obj
            campaign_template_obj.category = category_obj
            campaign_template_obj.status = template_status_obj
            campaign_template_obj.template_type = template_type_obj
            campaign_template_obj.bot = bot_obj
            campaign_template_obj.save()

        error_data["template_name_exist"] = template_name_exist_list
        error_data["template_language_exist_list"] = template_language_exist_list
        error_data["template_exist_rows"] = template_exist_rows
        # error_data["template_id_exist"] = template_id_exist_list
    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("In create_campaign_templates_from_file: %s at %s", str(e), str(exc_tb.tb_lineno), extra={
            'AppName': 'Campaign'})

        error_data["error"] = "Not able to create template. Please try again"

    return error_data


def get_user_details_from_batch(campaign_batch, CampaignFileAccessManagement):
    try:
        user_details = []
        header_map = {}
        key = campaign_batch.file_path.split('/')[-2]
        file_access_management_obj = CampaignFileAccessManagement.objects.filter(
            key=key)
        file_path = file_access_management_obj[0].file_path
        file_extension = file_path.split('.')[-1]

        file_path = settings.BASE_DIR + "/" + file_path

        if file_extension not in ['csv', 'psv']:
            ensure_element_tree(xlrd)
            workbook = xlrd.open_workbook(file_path)
            sheet = workbook.sheet_by_index(0)

            for col in range(0, sheet.ncols):
                header_map[sheet.cell_value(0, col)] = col

            for row in range(1, sheet.nrows):
                user_detail = []
                for col in range(0, sheet.ncols):
                    user_detail.append(sheet.cell_value(row, col))

                user_details.append(user_detail)
        else:
            if file_extension == "psv":
                delimiter = "|"
            else:
                delimiter = ","
            with open(file_path, 'r') as csv_sheet:
                csv_reader = csv.reader(csv_sheet, delimiter=delimiter)
                csv_header = next(csv_reader)
                is_pipe = check_pipe_delimiter_in_csv(csv_header)
                if is_pipe:
                    delimiter = "|"
                    psv_sheet = open(file_path, 'r+')
                    csv_reader = csv.reader(psv_sheet, delimiter=delimiter)
                    csv_header = next(csv_reader)
                for col in range(len(csv_header)):
                    header_map[csv_header[col].replace(',', '')] = col

                for row in csv_reader:
                    if not row:
                        continue
                    user_details.append(row)

                if is_pipe:
                    psv_sheet.close()

    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("get_mobile_no_from_batch Failed: %s at %s",
                     str(e), str(exc_tb.tb_lineno), extra={'AppName': 'Campaign'})

    return user_details, header_map


def get_language_code(template):
    try:
        language = template.language.title
        lang_code = re.findall(r'\((.*?)\)', language)

        if len(lang_code) != 0:
            return lang_code[-1]

        return 'en'
    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("get_language_code Failed: %s at %s",
                     str(e), str(exc_tb.tb_lineno), extra={'AppName': 'Campaign'})

        return 'en'


def get_unique_template_variable_body(template_body, variables):
    try:
        template_body_variables = re.findall(
            r'\{\{(.*?)\}\}', template_body)
        template_body_variables_list = list(
            dict.fromkeys(template_body_variables))
        variabe_mapping = dict(zip(template_body_variables_list, variables))

        template_body = re.sub(r'\{{.*?\}}', '{}', template_body)
        for body_variable in template_body_variables:
            template_body = template_body.replace(
                "{}", "{{" + variabe_mapping[body_variable] + "}}", 1)
    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("In get_unique_template_variable_body Failed: %s at %s",
                     str(e), str(exc_tb.tb_lineno), extra={'AppName': 'Campaign'})

    return template_body


def get_variable_details(variables, user_detail, header_map, dynamic_variable_url):
    try:
        variable_details = []
        unfilled_variables = set()
        for var_name in variables:
            col_num = header_map.get(var_name)
            if col_num == None:
                value = ''
            else:
                value = user_detail[col_num]

            if isinstance(value, float):
                if value == int(value):
                    value = str(int(value))
                else:
                    value = str(value)

            if dynamic_variable_url:
                dynamic_variable_url = dynamic_variable_url + '/'
                value = value.replace('https://' + dynamic_variable_url, '')

                value = value.replace('http://' + dynamic_variable_url, '')
                
                value = value.replace(dynamic_variable_url, '')

            if not value:
                unfilled_variables.add(var_name)
            variable_details.append(value)

        return variable_details, unfilled_variables
    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("get_variable_details Failed: %s at %s",
                     str(e), str(exc_tb.tb_lineno), extra={'AppName': 'Campaign'})

        return [], set()


def execute_send_rcs_campaign(campaign_id, schedule_obj, Campaign, CampaignFileAccessManagement, CampaignAudience, CampaignAudienceLog, CampaignAnalytics, CampaignTemplateVariable):
    try:
        campaign_obj = Campaign.objects.get(pk=int(campaign_id))

        logger.info("Campaign: %s started", campaign_obj.name, extra={
            'AppName': 'Campaign'})

        campaign_batch = campaign_obj.batch
        campaign_template = campaign_obj.campaign_template_rcs

        user_details, header_map = get_user_details_from_batch(
            campaign_batch, CampaignFileAccessManagement)

        if CampaignRCSDetailedAnalytics.objects.filter(campaign=campaign_obj).count() == 0:
            analytics_obj = CampaignRCSDetailedAnalytics.objects.create(
                campaign=campaign_obj,
                submitted=campaign_batch.total_audience,
                template=campaign_template,
                message_type=campaign_template.get_message_type_display(),
                bot=campaign_obj.bot)
        else:
            analytics_obj = CampaignRCSDetailedAnalytics.objects.filter(
                campaign=campaign_obj).first()

        tz = pytz.timezone(settings.TIME_ZONE)

        is_completed_successfully = True
        unsuccessful = 0
        analytics_obj.start_time = timezone.now().astimezone(tz)
        for user_detail in user_details:
            if user_detail[5] != 'True':
                continue

            user_detail[0] = get_valid_rcs_number(user_detail[0])
            audience_obj = CampaignAudience.objects.create(
                audience_id=int(user_detail[0]), channel=campaign_obj.channel, batch=campaign_batch, campaign=campaign_obj, record=json.dumps(user_detail))

            if campaign_template.message_type == '1':
                message_text, suggestion_chip_list = parse_text_message_type(
                    json.loads(campaign_template.template_metadata))
                message_text = messages.TextMessage(message_text)

            if campaign_template.message_type == '2':
                media_url, suggestion_chip_list = parse_media_message_type(
                    json.loads(campaign_template.template_metadata))
                message_text = messages.FileMessage(media_url)

            if campaign_template.message_type == '3':
                card_content, suggestion_chip_list = parse_card_message_type(
                    json.loads(campaign_template.template_metadata))
                if card_content != {}:
                    card_title = card_content["card_title"]
                    card_media_url = card_content["card_media_url"]
                    card_description = card_content["card_description"]
                    card_reply = card_content["card_reply"]

                    card_suggetions = append_suggestion_chip_to_cluster(
                        None, card_reply, messages)

                    message_text = messages.StandaloneCard('VERTICAL',
                                                           card_title,
                                                           card_description,
                                                           card_suggetions,
                                                           card_media_url,
                                                           None,
                                                           None,
                                                           'MEDIUM')

            if campaign_template.message_type == '4':
                carousel_content, suggestion_chip_list = parse_carousel_message_type(
                    json.loads(campaign_template.template_metadata))
                cards = []
                for carousel in carousel_content:
                    card_content, _ = parse_card_message_type(carousel)
                    if card_content != {}:
                        card_title = card_content["card_title"]
                        card_media_url = card_content["card_media_url"]
                        card_description = card_content["card_description"]
                        card_reply = card_content["card_reply"]

                        card_suggetions = append_suggestion_chip_to_cluster(
                            None, card_reply, messages)
                        cards.append(messages.CardContent(card_title,
                                                          card_description,
                                                          card_media_url,
                                                          'SHORT',
                                                          card_suggetions))

                message_text = messages.CarouselCard('MEDIUM', cards)

            # Send text message to the device
            cluster = messages.MessageCluster().append_message(message_text)
            cluster = append_suggestion_chip_to_cluster(
                cluster, suggestion_chip_list, messages)

            rcs_obj = RCSDetails.objects.filter(bot=campaign_obj.bot)

            if rcs_obj.exists():
                rcs_obj = rcs_obj.first()

                service_account_location = settings.BASE_DIR + "/" + rcs_obj.rcs_credentials_file_path

                response = cluster.send_to_msisdn(
                    user_detail[0], service_account_location)

                response = json.loads(response[1])
                audience_log_obj = CampaignAudienceLog.objects.create(
                    audience=audience_obj, campaign=campaign_obj)

            if 'name' in response:
                recepient_id = response["name"].split('/')[3]
                audience_log_obj.recepient_id = recepient_id
            else:
                unsuccessful += 1
                is_completed_successfully = False

            audience_log_obj.request = json.dumps({})
            audience_log_obj.response = json.dumps({})
            audience_log_obj.is_processed = True
            audience_log_obj.save()

            analytics_obj.end_time = timezone.now().astimezone(tz)
            analytics_obj.failed = unsuccessful
            analytics_obj.save()

        if is_completed_successfully and schedule_obj != None:
            schedule_obj.is_sent = True
            schedule_obj.save()

        if is_completed_successfully:
            campaign_obj.status = CAMPAIGN_COMPLETED
        else:
            campaign_obj.status = CAMPAIGN_PARTIALLY_COMPLETED
            if unsuccessful == campaign_obj.batch.total_audience_opted:
                campaign_obj.status = CAMPAIGN_FAILED

        campaign_obj.save()

    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("In execute_send_rcs_campaign: %s at %s", str(e), str(exc_tb.tb_lineno), extra={
            'AppName': 'Campaign'})

        campaign_obj.status = CAMPAIGN_PARTIALLY_COMPLETED

        campaign_obj.save()


def execute_send_campaign(campaign_wsp_config_meta, campaign_id, Campaign, CampaignFileAccessManagement, CampaignAudience, CampaignAudienceLog, CampaignAnalytics, CampaignTemplateVariable, test_data=dict()):
    try:
        campaign_obj = Campaign.objects.filter(pk=int(campaign_id)).first()
        is_test_campaign = campaign_wsp_config_meta.get('is_test_campaign', False)
        total_audience_to_test = sys.maxsize

        if is_test_campaign:
            campaign_obj.campaign_in_test = True
            campaign_obj.save(update_fields=['campaign_in_test'])
            total_audience_to_test = test_data.get('total_audience_to_test')
        else:
            campaign_obj.status = CAMPAIGN_IN_PROGRESS
            campaign_obj.start_datetime = datetime.now()
            campaign_obj.show_processed_datetime = True
            campaign_obj.save(update_fields=['status', 'start_datetime', 'show_processed_datetime'])
        logger.info("Campaign: %s started", campaign_obj.name, extra={
            'AppName': 'Campaign'})

        campaign_batch = campaign_obj.batch
        campaign_template = campaign_obj.campaign_template
        template_type = campaign_template.template_type.title
        lang_code = get_language_code(campaign_template)

        template_variable_obj = CampaignTemplateVariable.objects.filter(campaign=campaign_obj, template=campaign_template, batch=campaign_obj.batch).first()
        
        attachment_details = template_variable_obj.attachment_details

        if not isinstance(attachment_details, dict):
            attachment_details = json.loads(attachment_details)

        is_static = attachment_details.get('is_static', '')
        custom_attachment_url = attachment_details.get('custom_attachment_url', '').strip()
        fall_back_doc_name = attachment_details.get('fall_back_doc_name', '').strip()
        static_doc_name = attachment_details.get('static_doc_name', '').strip()
        user_details, header_map = get_user_details_from_batch(
            campaign_batch, CampaignFileAccessManagement)
        code, namespace, enable_queuing_system = campaign_wsp_config_meta[
            "code"], campaign_wsp_config_meta["namespace"], campaign_wsp_config_meta["enable_queuing_system"]
        if enable_queuing_system:
            aws_sqs, sqs_client = campaign_wsp_config_meta[
                "aws_sqs"], campaign_wsp_config_meta["sqs_client"]
            sqs_domain = f'{settings.EASYCHAT_HOST_URL}/campaign/lambda/push-message/'
        analytics_obj = CampaignAnalytics.objects.filter(campaign=campaign_obj)
        is_analytics_obj_exists = analytics_obj.exists()

        if is_test_campaign:
            if is_analytics_obj_exists:
                analytics_obj.update(total_tested=F(
                    'total_tested') + total_audience_to_test)
            else:
                CampaignAnalytics.objects.create(
                    campaign=campaign_obj, total_tested=total_audience_to_test)
        else:
            if is_analytics_obj_exists and analytics_obj.first().total_audience < 1 and analytics_obj.first().total_tested > 0:
                analytics_obj.update(
                    total_audience=campaign_batch.total_audience)
            elif not is_analytics_obj_exists:
                CampaignAnalytics.objects.create(
                    campaign=campaign_obj, total_audience=campaign_batch.total_audience)

        total_variables = template_variable_obj.total_variables
        dynamic_cta_url_variable = template_variable_obj.dynamic_cta_url_variable
        header_variable = template_variable_obj.header_variable

        if total_variables > 0:
            variables = json.loads(template_variable_obj.variables)
        else:
            variables = []

        if len(dynamic_cta_url_variable) > 0:
            dynamic_cta_url_variable = json.loads(dynamic_cta_url_variable)

        if len(header_variable) > 0:
            header_variable = json.loads(header_variable)

        template_metadata_json = json.loads(
            campaign_template.template_metadata)
        
        api_key = get_whatsapp_api_access_token(campaign_obj.bot)

        for user_detail in user_details:
            if is_test_campaign:
                if total_audience_to_test:
                    total_audience_to_test -= 1
                else:
                    break
            
            document_file_name = template_metadata_json.get(
                'document_file_name', CAMPAIGN_DOCUMENT_FILE_NAME)
            if not document_file_name.strip():
                document_file_name = CAMPAIGN_DOCUMENT_FILE_NAME

            variable_details, unfilled_variables = get_variable_details(
                variables, user_detail, header_map, None)

            header_variable_details, header_unfilled_variables = get_variable_details(
                header_variable, user_detail, header_map, None)

            unfilled_variables |= header_unfilled_variables
            cta_link = campaign_template.cta_link
            if cta_link:
                cta_link_variable = re.findall(r'\{\{(.*?)\}\}', cta_link)
                if cta_link_variable:
                    cta_link_variable = '{{' + cta_link_variable[0] + '}}'
                if cta_link and cta_link_variable == cta_link[-len(cta_link_variable):]:
                    cta_link = cta_link[-len(cta_link_variable):]
            dynamic_cta_url_variable_details, cta_unfilled_variables = get_variable_details(
                dynamic_cta_url_variable, user_detail, header_map, cta_link)

            unfilled_variables |= cta_unfilled_variables

            attachment_src = campaign_template.attachment_src

            attachment_src_variable = list()

            if is_static:
                if template_type != 'text' and custom_attachment_url:
                    attachment_src = custom_attachment_url
                if template_type == 'document' and static_doc_name:
                    document_file_name = static_doc_name
            else:
                dynamic_attactment_column = attachment_details.get('dynamic_attactment_column', '')
                if dynamic_attactment_column != 'none':
                    attachment_src_variable.append(dynamic_attactment_column)
                
                if template_type == 'document':
                    dynamic_doc_name_variable = list()
                    dynamic_doc_name_column = attachment_details.get('doc_name_options', '').strip()
                    if dynamic_doc_name_column != 'none':
                        dynamic_doc_name_variable.append(dynamic_doc_name_column)
                    dynamic_doc_name_variable_details, dynamic_doc_name_unfilled_variables = get_variable_details(
                        dynamic_doc_name_variable, user_detail, header_map, None)

                    dynamic_doc_name_variable_details = dynamic_doc_name_variable_details[0].strip() if dynamic_doc_name_variable_details else ''

                    if dynamic_doc_name_variable_details:
                        document_file_name = dynamic_doc_name_variable_details
                    elif not dynamic_doc_name_variable_details and fall_back_doc_name:
                        document_file_name = fall_back_doc_name

                if template_type != 'text':
                    attachment_src_variable_details, attachment_src_unfilled_variables = get_variable_details(
                        attachment_src_variable, user_detail, header_map, None)
                    
                    unfilled_variables |= attachment_src_unfilled_variables
                    if attachment_src_variable_details:
                        attachment_src_variable_details = attachment_src_variable_details[0].strip()
                    
                    if attachment_src_variable_details:
                        attachment_src = attachment_src_variable_details
                    elif not attachment_src_variable_details and custom_attachment_url:
                        attachment_src = custom_attachment_url
                    else:
                        attachment_src = ''

            unfilled_variables = ', '.join(unfilled_variables)
            phone_number = user_detail[0]
            phone_number = remo_html_from_string(str(phone_number))
            phone_number = remo_special_tag_from_string(phone_number)
            phone_number = validation_obj.removing_phone_non_digit_element(
                phone_number)
            phone_number = phone_number.strip()
            if phone_number[-2:] == ".0":
                phone_number = phone_number.replace('.0', '')
            user_detail.append(attachment_src)
            audience_obj = CampaignAudience.objects.create(
                audience_id=int(phone_number), channel=campaign_obj.channel, batch=campaign_batch, campaign=campaign_obj, record=json.dumps(user_detail))

            if (variables or dynamic_cta_url_variable or header_variable or attachment_src_variable) and unfilled_variables:
                audience_log_obj = CampaignAudienceLog.objects.create(
                    audience=audience_obj, campaign=campaign_obj, is_test=is_test_campaign)
                audience_log_obj.is_failed = True
                failure_response = {"errors": [{"code": 500, "title": "Failed to send message due to missing variable.",
                                                "details": f"Expected a value in the batch file column/s {unfilled_variables} selected for the variable/s in the message. Please fill in the required data for this user"}]}
                audience_log_obj.request = json.dumps({})
                audience_log_obj.response = json.dumps(failure_response)
                audience_log_obj.is_processed = True
                audience_log_obj.save(
                    update_fields=["request", "response", "is_failed", "is_processed"])
                continue

            if is_livechat_connected(phone_number, campaign_obj.bot):
                audience_log_obj = CampaignAudienceLog.objects.create(
                    audience=audience_obj, campaign=campaign_obj, is_test=is_test_campaign)
                livechat_active_session_failure_response(audience_log_obj)
                continue

            parameter = {
                'mobile_number': phone_number,
                'template': {
                    'name': campaign_template.template_name,
                    'type': campaign_template.template_type.title,
                    'language': lang_code,
                    'link': attachment_src,
                    'cta_text': campaign_template.cta_text,
                    'cta_link': campaign_template.cta_link,
                },
                'user_details': user_detail,
                'variables': variable_details,
                'header_variable': header_variable_details,
                'dynamic_cta_url_variable': dynamic_cta_url_variable_details,
                'type_of_first_cta_btton': template_metadata_json.get('type_of_first_cta_btton'),
                'document_file_name': document_file_name,
                'namespace': namespace,
                'api_key': api_key,
            }

            logger.info("parameter: %s", str(parameter), extra={
                'AppName': 'Campaign'})

            if enable_queuing_system:
                sqs_packet = {
                    "url": sqs_domain,
                    "campaign_id": campaign_id,
                    "bot_wsp_id": campaign_wsp_config_meta["bot_wsp_id"],
                    "audience_id": audience_obj.pk,
                    "parameter": parameter,
                    "event_name": "SEND_MESSAGE_EVENT"
                }
                send_message_into_campaign_sqs(
                    json.dumps(sqs_packet), sqs_client, aws_sqs)

            else:
                audience_log_obj = CampaignAudienceLog.objects.create(
                    audience=audience_obj, campaign=campaign_obj, is_test=is_test_campaign)
                processor_check_dictionary = {'open': open_file}
                exec(str(code), processor_check_dictionary)

                response = processor_check_dictionary['f'](
                    json.dumps(parameter))

                logger.info("response: %s", str(response), extra={
                    'AppName': 'Campaign'})

                if response.get('response') and 'request_id' in response['response']:
                    audience_log_obj.recepient_id = response['response']['request_id']
                else:
                    audience_log_obj.is_failed = True

                if response.get('request') and response.get('response'):
                    audience_log_obj.request = json.dumps(response['request'])
                    audience_log_obj.response = json.dumps(
                        response['response'])

                audience_log_obj.is_processed = True
                audience_log_obj.save(
                    update_fields=["recepient_id", "request", "response", "is_failed", "is_processed"])

        if is_test_campaign:
            campaign_obj.campaign_in_test = False
            campaign_obj.save(update_fields=['campaign_in_test'])

    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("In execute_send_campaign: %s at %s", str(e), str(exc_tb.tb_lineno), extra={
            'AppName': 'Campaign'})


def check_for_system_commands(code, CampaignConfig):
    try:
        config_obj = CampaignConfig.objects.all()[0]
        system_commands = json.loads(
            config_obj.system_commands.replace("'", '"'))

        for command in system_commands:
            if command in code:
                return True
    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("check_for_system_commands: %s %s", str(e), str(exc_tb.tb_lineno), extra={
            'AppName': 'Campaign'})

    return False


def ensure_element_tree(xlrd):
    try:
        xlrd.xlsx.ensure_elementtree_imported(False, None)
        xlrd.xlsx.Element_has_iter = True
    except Exception:
        pass


def get_schedule_metadata(time_slots, batch_obj_pk):
    try:
        meta_data = []
        for time in time_slots:
            uid = uuid.uuid4()
            obj = {
                "time": time,
                "uid": str(uid),
                "batch_pk": batch_obj_pk,
            }
            meta_data.append(obj)

        return json.dumps(meta_data)
    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("get_schedule_metadata: %s %s", str(e), str(exc_tb.tb_lineno), extra={
            'AppName': 'Campaign'})

    return meta_data


def create_objects_related_to_schedule(start_date, end_date, choice_selected, metadata, campaign_schedule_obj, CampaignScheduleObject, days, upto_date=None, days_limit="", batch_pk=None):
    try:
        for meta_obj in metadata:
            delta = -1
            if upto_date != None:
                delta = upto_date - start_date
                delta = delta.days
                delta = delta + 1

            time = datetime.strptime(meta_obj["time"], '%I:%M %p').time()
            uid = uuid.UUID(meta_obj["uid"])
            batch_pk = meta_obj.get("batch_pk", batch_pk) 
            edited_on = meta_obj.get("edited_on")
            edited_on = datetime.strptime(edited_on, "%d %b %Y, %I:%M %p") if edited_on else edited_on
            if meta_obj["time"] != "":
                if choice_selected == "does_not_repeat":
                    create_schedule_obj(
                        campaign_schedule_obj, start_date, time, uid, CampaignScheduleObject, batch_pk, edited_on)
                elif choice_selected == "daily":
                    if days_limit != "":
                        delta = days_limit
                    if delta == -1:
                        delta = 90
                    thread = threading.Thread(target=create_schedule_obj_daily, args=(
                        campaign_schedule_obj, start_date, time, uid, CampaignScheduleObject, delta, batch_pk, edited_on), daemon=True)
                    thread.start()
                    logger.info("Threading started...for creating daily objects", extra={
                        'AppName': 'Campaign', 'user_id': 'None', 'source': 'None', 'channel': 'None', 'bot_id': ''})
                elif choice_selected == "weekly":
                    if delta >= 1:
                        delta = delta // 7
                        delta += 1
                        if days_limit != "":
                            delta = days_limit
                    elif delta == -1:
                        delta = 60
                    thread = threading.Thread(target=create_schedule_obj_weekly, args=(
                        campaign_schedule_obj, start_date, time, uid, CampaignScheduleObject, days, delta, batch_pk, edited_on), daemon=True)
                    thread.start()
                    logger.info("Threading started...for creating weekly objects", extra={
                        'AppName': 'Campaign', 'user_id': 'None', 'source': 'None', 'channel': 'None', 'bot_id': ''})
                elif choice_selected == "monthly":
                    if delta >= 1:
                        delta = delta // 30
                        delta += 1
                        if days_limit != "":
                            delta = days_limit
                    elif delta == -1:
                        delta = 60
                    thread = threading.Thread(target=create_schedule_obj_monthly, args=(
                        campaign_schedule_obj, start_date, time, uid, CampaignScheduleObject, days, delta, batch_pk, edited_on), daemon=True)
                    thread.start()
                    logger.info("Threading started...for creating weekly objects", extra={
                        'AppName': 'Campaign', 'user_id': 'None', 'source': 'None', 'channel': 'None', 'bot_id': ''})
                elif choice_selected == "annually":
                    if delta >= 365:
                        delta = delta // 365
                        delta += 1
                        if days_limit != "":
                            delta = days_limit / 365
                            delta = int(math.ceil(delta))
                            delta = delta + 1
                    elif delta == -1:
                        delta = 60
                    thread = threading.Thread(target=create_schedule_obj_annually, args=(
                        campaign_schedule_obj, start_date, time, uid, CampaignScheduleObject, delta, batch_pk, edited_on), daemon=True)
                    thread.start()
                    logger.info("Threading started...for creating weekly objects", extra={
                        'AppName': 'Campaign', 'user_id': 'None', 'source': 'None', 'channel': 'None', 'bot_id': ''})
                elif choice_selected == "weekend_or_weekday":
                    if delta >= 1:
                        if days_limit != "":
                            delta = days_limit
                            delta += 1
                    elif delta == -1:
                        delta = 90
                    thread = threading.Thread(target=create_schedule_obj_weekday_or_weekend, args=(
                        campaign_schedule_obj, start_date, time, uid, CampaignScheduleObject, delta, batch_pk, edited_on), daemon=True)
                    thread.start()
                    logger.info("Threading started...for creating weekly objects", extra={
                        'AppName': 'Campaign', 'user_id': 'None', 'source': 'None', 'channel': 'None', 'bot_id': ''})
                elif choice_selected == "custom":
                    if delta >= 1:
                        if days_limit != "":
                            delta = days_limit
                            delta += 1
                    elif delta == -1:
                        delta = 90
                    thread = threading.Thread(target=create_schedule_obj_custom, args=(
                        campaign_schedule_obj, start_date, time, uid, CampaignScheduleObject, days, end_date, delta, batch_pk, edited_on), daemon=True)
                    thread.start()
                    logger.info("Threading started...for creating weekly objects", extra={
                        'AppName': 'Campaign', 'user_id': 'None', 'source': 'None', 'channel': 'None', 'bot_id': ''})

    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("create_objects_related_to_schedule: %s %s", str(e), str(exc_tb.tb_lineno), extra={
            'AppName': 'Campaign'})


def create_schedule_obj(campaign_schedule, date, time, uid, CampaignScheduleObject, batch_pk, edited_on):
    try:
        current_datetime = datetime.now()
        date_and_time = datetime.combine(date, time)
        if current_datetime < date_and_time:
            create_schedule_obj = CampaignScheduleObject.objects.filter(
                campaign=campaign_schedule.campaign,
                date=date,
                time=time,
            ).first()
            if not create_schedule_obj:
                create_schedule_obj = CampaignScheduleObject.objects.create(
                    campaign_schedule=campaign_schedule,
                    date=date,
                    time=time,
                    uid=uid,
                    campaign=campaign_schedule.campaign,
                    channel=campaign_schedule.campaign.channel,
                    campaign_batch_id=batch_pk,
                    edited_on=edited_on,
                )
                campaign_schedule.campaign.status = CAMPAIGN_SCHEDULED
                campaign_schedule.campaign.save()
    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("create_schedule_obj: %s %s", str(e), str(exc_tb.tb_lineno), extra={
            'AppName': 'Campaign'})


def create_schedule_obj_daily(campaign_schedule, date, time, uid, CampaignScheduleObject, days_range, batch_pk, edited_on):
    try:
        # create 3 months data at once daily
        start_date = date
        for day in range(days_range):
            if start_date != None:
                date = start_date + timedelta(days=day)
                create_schedule_obj(campaign_schedule, date,
                                    time, uid, CampaignScheduleObject, batch_pk, edited_on)

    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("create_schedule_obj_daily: %s %s", str(e), str(exc_tb.tb_lineno), extra={
            'AppName': 'Campaign'})


def create_schedule_obj_weekly(campaign_schedule, date, time, uid, CampaignScheduleObject, days, days_range, batch_pk, edited_on):
    try:
        # create 3 months data at once weekly
        start_date = date
        for day in range(days_range):
            date = start_date + timedelta(days=7 * day)
            create_schedule_obj(campaign_schedule, date,
                                time, uid, CampaignScheduleObject, batch_pk, edited_on)
    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("create_schedule_obj_weekly: %s %s", str(e), str(exc_tb.tb_lineno), extra={
            'AppName': 'Campaign'})


def create_schedule_obj_monthly(campaign_schedule, date, time, uid, CampaignScheduleObject, days, days_range, batch_pk, edited_on):
    try:
        # create 3 months data at once monthly
        start_date = date
        for day in range(days_range):
            date = start_date + relativedelta.relativedelta(months=day)
            date = get_day_with_index(date, days)
            create_schedule_obj(campaign_schedule, date,
                                time, uid, CampaignScheduleObject, batch_pk, edited_on)
    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("create_schedule_obj_monthly: %s %s", str(e), str(exc_tb.tb_lineno), extra={
            'AppName': 'Campaign'})


def create_schedule_obj_annually(campaign_schedule, date, time, uid, CampaignScheduleObject, days_range, batch_pk, edited_on):
    try:
        # create 1 year data at once yearly
        start_date = date
        for year in range(days_range):
            date = start_date + relativedelta.relativedelta(years=year)
            create_schedule_obj(campaign_schedule, date,
                                time, uid, CampaignScheduleObject, batch_pk, edited_on)
    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("create_schedule_obj_annually: %s %s", str(e), str(exc_tb.tb_lineno), extra={
            'AppName': 'Campaign'})


def create_schedule_obj_weekday_or_weekend(campaign_schedule, date, time, uid, CampaignScheduleObject, days_range, batch_pk, edited_on):
    try:
        # create 3 months data at once weekly
        start_date = date
        all_days = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
        weekday = ["Mon", "Tue", "Wed", "Thu", "Fri"]
        weekend = ["Sat", "Sun"]
        current_day = all_days[start_date.weekday()]
        # week
        for day in range(days_range):
            date = start_date + timedelta(days=day)
            if current_day == "Sat" or current_day == "Sun":
                if all_days[date.weekday()] in weekend:
                    create_schedule_obj(
                        campaign_schedule, date, time, uid, CampaignScheduleObject, batch_pk, edited_on)
            else:
                if all_days[date.weekday()] in weekday:
                    create_schedule_obj(
                        campaign_schedule, date, time, uid, CampaignScheduleObject, batch_pk, edited_on)

    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("create_schedule_obj_weekday_or_weekend: %s %s", str(e), str(exc_tb.tb_lineno), extra={
            'AppName': 'Campaign'})


def create_schedule_obj_custom(campaign_schedule, date, time, uid, CampaignScheduleObject, days, end_date, days_range, batch_pk, edited_on):
    try:
        # create custom which ends never
        all_days = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
        start_date = date

        if end_date == None:
            delta = days_range
        else:
            delta = end_date - start_date
            delta = delta.days
        for day in range(delta):
            date = start_date + timedelta(days=day)
            current_day = all_days[date.weekday()]
            try:
                if days.index(current_day) >= 0:
                    create_schedule_obj(
                        campaign_schedule, date, time, uid, CampaignScheduleObject, batch_pk, edited_on)
            except ValueError:
                pass

    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("create_schedule_obj_custom: %s %s", str(e), str(exc_tb.tb_lineno), extra={
            'AppName': 'Campaign'})


def get_day_with_index(date, days):
    try:
        all_days = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
        day = days[0]
        day_of_week = day.split("-")[0]
        index = day.split("-")[1]

        occurence_in_list = all_days.index(day_of_week)
        cal = calendar.Calendar(firstweekday=calendar.SUNDAY)

        year = date.year
        month = date.month

        monthcal = cal.monthdatescalendar(year, month)
        days_list = [day for week in monthcal for day in week if day.weekday(
        ) == occurence_in_list and day.month == month]

        avail_days = ["First", "Second", "Third", "Fourth", "Last"]
        occurence_in_list = avail_days.index(index)
        if index == "Last":
            return days_list[len(days_list) - 1]
        else:
            return days_list[occurence_in_list]
    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("get_day_with_index: %s %s", str(e), str(exc_tb.tb_lineno), extra={
            'AppName': 'Campaign'})


def get_schedules_list_obj(campaign_schedules):
    try:
        schedules = []
        for schedule_obj in campaign_schedules:
            last_modified = schedule_obj.edited_on
            temp_schedule = {
                "campaign_name": schedule_obj.campaign_schedule.campaign.name,
                "trigger_date": str(schedule_obj.date.strftime("%d-%b-%Y")),
                "trigger_time": str(schedule_obj.time.strftime("%I:%M %p")),
                "created_on": str((schedule_obj.campaign_schedule.created_at + timedelta(hours=5, minutes=30)).strftime("%d %b %Y, %I:%M %p")),
                "schedule_pk": schedule_obj.campaign_schedule.pk,
                "current_slot_pk": schedule_obj.pk,
                "uid": str(schedule_obj.uid),
                "batch_name": schedule_obj.campaign_batch.batch_name,
                "last_modified": str((last_modified + timedelta(hours=5, minutes=30)).strftime("%d %b %Y, %I:%M %p")) if last_modified else None,
            }
            schedules.append(temp_schedule)
        return schedules
    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("get_schedules_list_obj: %s %s", str(e), str(exc_tb.tb_lineno), extra={
            'AppName': 'Campaign'})

    return schedules


def create_remaining_schedule_objects(campaign_schedule, upto_date, CampaignScheduleObject):
    try:
        date_format = "%Y-%m-%d"
        start_date = campaign_schedule.updated_upto
        end_date = campaign_schedule.end_date
        if start_date != None:
            start_date = datetime.strptime(
                str(start_date), date_format).date()
        if upto_date != None:
            upto_date = datetime.strptime(
                str(upto_date), date_format).date()
        if end_date != None:
            end_date = datetime.strptime(
                str(end_date), date_format).date()
        create_remaining = True
        if campaign_schedule.choices == "custom" and end_date != None:
            create_remaining = False
        if campaign_schedule.choices != "does_not_repeat" and create_remaining:
            create_objects_related_to_schedule(start_date, end_date, campaign_schedule.choices, json.loads(
                campaign_schedule.metadata), campaign_schedule, CampaignScheduleObject, json.loads(campaign_schedule.days), upto_date)
    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("create_remaining_schedule_objects: %s %s", str(e), str(exc_tb.tb_lineno), extra={
            'AppName': 'Campaign'})


def create_new_added_slot_schedule_objects(campaign_schedule, CampaignScheduleObject, current_obj):
    try:
        date_format = "%Y-%m-%d"

        if campaign_schedule.start_date != None:
            start_date = datetime.strptime(
                str(campaign_schedule.start_date), date_format).date()
        if campaign_schedule.updated_upto != None:
            upto_date = datetime.strptime(
                str(campaign_schedule.updated_upto), date_format).date()

        end_date = None
        delta = upto_date - start_date
        delta = delta.days
        delta = delta + 1
        if delta >= 0:
            if campaign_schedule.choices == "weekly":
                delta = delta / 7
                delta = int(math.ceil(delta))
            elif campaign_schedule.choices == "monthly":
                delta = delta / 30
                delta = int(math.ceil(delta))
                delta = delta + 1
            create_objects_related_to_schedule(start_date, end_date, campaign_schedule.choices, json.loads(
                campaign_schedule.metadata), campaign_schedule, CampaignScheduleObject, json.loads(campaign_schedule.days), upto_date, delta)

    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("create_new_added_slot_schedule_objects: %s %s", str(e), str(exc_tb.tb_lineno), extra={
            'AppName': 'Campaign'})


def check_and_create_trigger_setting(campaign_obj, CampaignVoiceBotSetting, VoiceBotRetrySetting):
    try:
        voice_bot_obj = CampaignVoiceBotSetting.objects.filter(
            campaign=campaign_obj)

        if not voice_bot_obj:
            voice_bot_obj = CampaignVoiceBotSetting.objects.create(
                campaign=campaign_obj)

            retry_setting = VoiceBotRetrySetting.objects.create()

            voice_bot_obj.retry_setting = retry_setting
            voice_bot_obj.save()
        else:
            voice_bot_obj = voice_bot_obj.first()

        return voice_bot_obj
    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("check_and_create_trigger_setting: %s %s", str(e), str(exc_tb.tb_lineno), extra={
            'AppName': 'Campaign'})


def parse_trigger_settings(voice_bot_obj):
    try:
        retry_setting = voice_bot_obj.retry_setting

        if voice_bot_obj.is_saved:
            start_date = voice_bot_obj.start_date.strftime('%Y/%m/%d')
            start_time = voice_bot_obj.start_time.strftime(TIME_FORMAT)
            end_date = voice_bot_obj.end_date.strftime('%Y/%m/%d')
            end_time = voice_bot_obj.end_time.strftime(TIME_FORMAT)
        else:
            start_date = datetime.now().date().strftime('%Y/%m/%d')
            start_time = (datetime.now() + timedelta(minutes=10)
                          ).time().strftime(TIME_FORMAT)
            end_date = (datetime.now() + timedelta(days=30)
                        ).date().strftime('%Y/%m/%d')
            end_time = None

        on_status = []
        if retry_setting.is_busy_enabled:
            on_status.append('Busy')

        if retry_setting.is_no_answer_enabled:
            on_status.append('No Answer')

        if retry_setting.is_failed_enabled:
            on_status.append('Failed')

        trigger_setting = {
            'caller_id': voice_bot_obj.caller_id,
            'start_date': start_date,
            'end_date': end_date,
            'start_time': start_time,
            'end_time': end_time,
            'retry_mechanism': retry_setting.mechanism,
            'no_of_retries': retry_setting.no_of_retries,
            'retry_interval': retry_setting.retry_interval,
            'is_busy_enabled': retry_setting.is_busy_enabled,
            'is_no_answer_enabled': retry_setting.is_no_answer_enabled,
            'is_failed_enabled': retry_setting.is_failed_enabled,
            'on_status': on_status,
        }

        return trigger_setting
    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("parse_trigger_settings: %s %s", str(e), str(exc_tb.tb_lineno), extra={
            'AppName': 'Campaign'})


def get_mob_numbers_from_user_details(user_details):
    try:
        mob_numbers = []
        for user_detail in user_details:
            phone_number = user_detail[0]
            phone_number = remo_html_from_string(str(phone_number))
            phone_number = remo_special_tag_from_string(phone_number)
            phone_number = validation_obj.removing_phone_non_digit_element(
                phone_number)
            phone_number = phone_number.strip()
            if phone_number[-2:] == ".0":
                phone_number = phone_number.replace('.0', '')

            mob_numbers.append(str(int(phone_number)))

        return mob_numbers

    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("get_mob_numbers_from_user_details: %s %s", str(e), str(exc_tb.tb_lineno), extra={
            'AppName': 'Campaign'})

    return []


def check_and_create_voice_bot_api_obj(campaign_obj, CampaignVoiceBotAPI):
    try:
        api_obj = CampaignVoiceBotAPI.objects.filter(campaign=campaign_obj)

        if api_obj:
            return api_obj.first()

        api_obj = CampaignVoiceBotAPI.objects.create(campaign=campaign_obj)

        return api_obj

    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("get_mob_numbers_from_user_details: %s %s", str(e), str(exc_tb.tb_lineno), extra={
            'AppName': 'Campaign'})

    return None


def get_app_id_from_caller_id(caller_id, VoiceBotCallerID):
    try:
        caller_id_obj = VoiceBotCallerID.objects.filter(caller_id=caller_id)

        if caller_id_obj:
            caller_id_obj = caller_id_obj.first()

            return caller_id_obj.app_id

    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("get_app_id_from_caller_id: %s %s", str(e), str(exc_tb.tb_lineno), extra={
            'AppName': 'Campaign'})

    return None


def get_campaign_rcs_analytics(campaign_rcs_analytics_objs):
    try:
        submitted = campaign_rcs_analytics_objs.aggregate(Sum('submitted'))[
            'submitted__sum']
        submitted = 0 if submitted == None else submitted

        sent = campaign_rcs_analytics_objs.aggregate(Sum('sent'))[
            'sent__sum']
        sent = 0 if sent == None else sent

        delivered = campaign_rcs_analytics_objs.aggregate(Sum('delivered'))[
            'delivered__sum']
        delivered = 0 if delivered == None else delivered

        read = campaign_rcs_analytics_objs.aggregate(Sum('read'))[
            'read__sum']
        read = 0 if read == None else read

        failed = campaign_rcs_analytics_objs.aggregate(Sum('failed'))[
            'failed__sum']
        failed = 0 if failed == None else failed

        replied = campaign_rcs_analytics_objs.aggregate(Sum('replied'))[
            'replied__sum']
        replied = 0 if replied == None else replied

        return submitted, sent, delivered, read, failed, replied

    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("get_campaign_rcs_analytics: %s %s", str(e), str(exc_tb.tb_lineno), extra={
            'AppName': 'Campaign'})

    return 0, 0, 0, 0, 0, 0


def get_campaign_voice_analytics(campaign_voice_analytics_objs):
    try:
        total_calls_created = campaign_voice_analytics_objs.aggregate(Sum('call_created'))[
            'call_created__sum']
        total_calls_created = 0 if total_calls_created == None else total_calls_created

        total_calls_scheduled = campaign_voice_analytics_objs.aggregate(Sum('call_scheduled'))[
            'call_scheduled__sum']
        total_calls_scheduled = 0 if total_calls_scheduled == None else total_calls_scheduled

        total_calls_initiated = campaign_voice_analytics_objs.aggregate(Sum('call_initiated'))[
            'call_initiated__sum']
        total_calls_initiated = 0 if total_calls_initiated == None else total_calls_initiated

        total_calls_in_progress = campaign_voice_analytics_objs.aggregate(Sum('call_in_progress'))[
            'call_in_progress__sum']
        total_calls_in_progress = 0 if total_calls_in_progress == None else total_calls_in_progress

        total_calls_retry = campaign_voice_analytics_objs.aggregate(Sum('call_retry'))[
            'call_retry__sum']
        total_calls_retry = 0 if total_calls_retry == None else total_calls_retry

        total_calls_retrying = campaign_voice_analytics_objs.aggregate(Sum('call_retrying'))[
            'call_retrying__sum']
        total_calls_retrying = 0 if total_calls_retrying == None else total_calls_retrying

        total_calls_retry += total_calls_retrying

        total_calls_failed = campaign_voice_analytics_objs.aggregate(Sum('call_failed'))[
            'call_failed__sum']
        total_calls_failed = 0 if total_calls_failed == None else total_calls_failed

        total_calls_failed_dnd = campaign_voice_analytics_objs.aggregate(Sum('call_failed_dnd'))[
            'call_failed_dnd__sum']
        total_calls_failed_dnd = 0 if total_calls_failed_dnd == None else total_calls_failed_dnd

        total_calls_failed += total_calls_failed_dnd

        total_calls_completed = campaign_voice_analytics_objs.aggregate(Sum('call_completed'))[
            'call_completed__sum']
        total_calls_completed = 0 if total_calls_completed == None else total_calls_completed

        return total_calls_created, total_calls_scheduled, total_calls_initiated, total_calls_in_progress, total_calls_retry, total_calls_failed, total_calls_completed

    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("get_campaign_voice_analytics: %s %s", str(e), str(exc_tb.tb_lineno), extra={
            'AppName': 'Campaign'})

    return 0, 0, 0, 0, 0, 0, 0


def get_campaign_whatsapp_analytics(campaign_analytics_objs):
    try:
        whatsapp_analytics = dict()

        total_messages_sent = campaign_analytics_objs.aggregate(Sum('message_sent'))[
            'message_sent__sum']
        whatsapp_analytics['total_messages_sent'] = 0 if total_messages_sent == None else total_messages_sent

        total_messages_read = campaign_analytics_objs.aggregate(Sum('message_read'))[
            'message_read__sum']
        whatsapp_analytics['total_messages_read'] = 0 if total_messages_read == None else total_messages_read

        total_messages_delivered = campaign_analytics_objs.aggregate(
            Sum('message_delivered'))['message_delivered__sum']
        whatsapp_analytics['total_messages_delivered'] = 0 if total_messages_delivered == None else total_messages_delivered

        total_messages_unsuccessful = campaign_analytics_objs.aggregate(
            Sum('message_unsuccessful'))['message_unsuccessful__sum']
        whatsapp_analytics['total_messages_unsuccessful'] = 0 if total_messages_unsuccessful == None else total_messages_unsuccessful

        total_messages_replied = campaign_analytics_objs.aggregate(
            Sum('message_replied'))['message_replied__sum']
        whatsapp_analytics['total_messages_replied'] = 0 if total_messages_replied == None else total_messages_replied

        total_messages_processed = campaign_analytics_objs.aggregate(
            Sum('message_processed'))['message_processed__sum']
        whatsapp_analytics['total_messages_processed'] = 0 if total_messages_processed == None else total_messages_processed

        test_message_sent = campaign_analytics_objs.aggregate(
            Sum('test_message_sent'))['test_message_sent__sum']
        whatsapp_analytics['test_message_sent'] = 0 if test_message_sent == None else test_message_sent

        test_message_unsuccessful = campaign_analytics_objs.aggregate(
            Sum('test_message_unsuccessful'))['test_message_unsuccessful__sum']
        whatsapp_analytics['test_message_unsuccessful'] = 0 if test_message_unsuccessful == None else test_message_unsuccessful

    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("get_campaign_voice_analytics: %s %s", str(e), str(exc_tb.tb_lineno), extra={
            'AppName': 'Campaign'})

    return whatsapp_analytics


def check_and_create_call_details_obj(campaign_obj, call_sid, CampaignVoiceBotDetailedAnalytics):
    try:
        call_details_obj = CampaignVoiceBotDetailedAnalytics.objects.filter(
            campaign=campaign_obj, call_sid=call_sid)

        if call_details_obj:
            return call_details_obj.first()

        call_details_obj = CampaignVoiceBotDetailedAnalytics.objects.create(
            campaign=campaign_obj, call_sid=call_sid)

        return call_details_obj

    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("check_and_create_call_details_obj: %s %s", str(e), str(exc_tb.tb_lineno), extra={
            'AppName': 'Campaign'})


def check_element_in_data(data, element):
    try:
        if element in data:
            return True

    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("check_element_in_data: %s %s", str(e), str(exc_tb.tb_lineno), extra={
            'AppName': 'Campaign'})

    return False


def add_rcs_enabled_column(file_path, bot_obj, RCSDetails):
    try:
        file_extension = file_path.split('.')[-1]
        file_path = settings.BASE_DIR + '/' + file_path

        if file_extension not in ['csv', 'psv']:
            ensure_element_tree(xlrd)
            wb = xlrd.open_workbook(file_path)
            sheet = wb.sheet_by_index(0)
            if file_extension != 'xls':
                wb = openpyxl.load_workbook(file_path)
                ws = wb["Sheet1"]
            else:
                wb = openpyxl.Workbook()
                ws = wb.create_sheet()
                for row in range(sheet.nrows):
                    for col in range(sheet.ncols):
                        ws.cell(row=row + 1, column=col + 1).value = sheet.cell_value(row, col)

            msdins = sheet.col_values(0, 1)
        else:
            if file_extension == "psv":
                delimiter = "|"
            else:
                delimiter = ","
            with open(file_path, 'r') as csv_sheet:
                csv_reader = csv.reader(csv_sheet, delimiter=delimiter)
                csv_header = next(csv_reader)
                row_data = [row for row in csv_reader]
                msdins = [row_value[0] for row_value in row_data]

            csv_updated_sheet = open(file_path, 'w')
            csv_writer = csv.writer(csv_updated_sheet, delimiter=delimiter)

        rcs_obj = RCSDetails.objects.filter(bot=bot_obj)
        rcs_enabled_user_list = []
        if rcs_obj.exists() and rcs_obj.first().rcs_credentials_file_path.strip() != "":
            rcs_obj = rcs_obj.first()
            service_account_location = rcs_obj.rcs_credentials_file_path
            total_audience_length = len(msdins)
            for index in range(total_audience_length):
                msdins[index] = get_valid_rcs_number(msdins[index])

            output_msdins = [msdins[itr:itr + 10000]
                             for itr in range(0, total_audience_length, 10000)]
            try:
                for iterator in output_msdins:
                    content = rbm_service.make_batch_cap_request(
                        iterator, service_account_location)
                    if "reachableUsers" in content:
                        rcs_enabled_user_list += content["reachableUsers"]
            except Exception as e:
                exc_type, exc_obj, exc_tb = sys.exc_info()
                logger.error("add_rcs_enabled_column: %s %s", str(e), str(exc_tb.tb_lineno), extra={
                    'AppName': 'Campaign'})
                return 0, 401, "Unable to upload file, please make sure at least one valid phone number with rcs enabled is provided."
            st = set(rcs_enabled_user_list)
            indices = [i for i, e in enumerate(msdins) if e in st]
            rcs_enabled_user_exists = False
            if file_extension != 'csv':
                ws.cell(row=1, column=6).value = "RCS Enabled (True/False)"
                max_row = sheet.nrows
            else:
                csv_header.append("RCS Enabled (True/False)")
                csv_writer.writerow(csv_header)
                max_row = len(row_data) + 1

            for iterator in range(0, max_row - 1):
                update_index = iterator + 2
                if iterator in indices:
                    rcs_enabled_user_exists = True
                    if file_extension != 'csv':
                        ws.cell(row=update_index, column=6).value = str(True)
                    else:
                        row_data[iterator].append(str(True))
                        csv_writer.writerow(row_data[iterator])
                else:
                    if file_extension != 'csv':
                        ws.cell(row=update_index, column=6).value = str(False)
                    else:
                        row_data[iterator].append(str(False))
                        csv_writer.writerow(row_data[iterator])

            if file_extension != 'csv':
                wb.save(file_path)
            else:
                csv_updated_sheet.close()
            if rcs_enabled_user_exists == False:
                return 0, 403, "Add atleast 1 RCS Enabled user"
            else:
                return len(indices), 200, "File uploaded successfully"
        else:
            status_message = "Please setup RCS channel for the bot properly."
            return 0, 401, status_message

    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("add_rcs_enabled_column: %s %s", str(e), str(exc_tb.tb_lineno), extra={
            'AppName': 'Campaign'})

        return 0, 401, "Unable to upload file"


def get_valid_rcs_number(number):
    try:
        if number[-2:] == ".0":
            number = number.replace('.0', '')
        if number[0] != "+":
            number = '+' + number

    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("Error FileAccess %s at %s",
                     str(e), str(exc_tb.tb_lineno), extra={'AppName': 'CampaignApp'})

    return number


def FileAccess(request, file_key, file_name):
    comes_from_excel = False
    if "source" in request.GET:
        comes_from_excel = True
    if not request.user.is_authenticated and comes_from_excel == False:
        return HttpResponse("Invalid Request")

    file_access_management_obj = None
    try:
        file_access_management_obj = CampaignFileAccessManagement.objects.get(
            key=file_key)

        if file_access_management_obj.bot:
            if request.user not in file_access_management_obj.bot.users.all():
                return HttpResponse("You are not authorized to download this file.")
      
        path_to_file = file_access_management_obj.file_path

        filename = path_to_file.split("/")[-1]

        path_to_file = settings.BASE_DIR + "/" + path_to_file

        mime_type, _ = mimetypes.guess_type(path_to_file)
        if os.path.exists(path_to_file):
            with open(path_to_file, 'rb') as fh:
                response = HttpResponse(
                    fh.read(), status=200, content_type=mime_type)
                response['Content-Disposition'] = 'attachment; filename=%s' % smart_str(
                    str(filename))
                # response['X-Sendfile'] = smart_str(path_to_file)
                # response['X-Accel-Redirect'] = path_to_file
                return response

        return HttpResponse(status=401)

    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("Error FileAccess %s at %s",
                     str(e), str(exc_tb.tb_lineno), extra={'AppName': 'CampaignApp'})
    return HttpResponse(status=404)


def check_and_create_config_obj(bot_obj, CampaignConfig):
    try:
        config_obj = CampaignConfig.objects.filter(bot=bot_obj).first()

        if config_obj:
            return config_obj

        config_obj = CampaignConfig.objects.create(bot=bot_obj)

        return config_obj
    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("check_and_create_config_obj: %s %s", str(e), str(exc_tb.tb_lineno), extra={
            'AppName': 'Campaign'})

        config_obj = CampaignConfig.objects.create()

        return config_obj


def check_and_create_campaign_api(channel, campaign_obj, CampaignAPI, CampaignVoiceBotAPI):
    try:
        if channel.value == 'voicebot':
            check_and_create_voice_bot_api_obj(
                campaign_obj, CampaignVoiceBotAPI)
        else:
            if CampaignAPI.objects.filter(campaign=campaign_obj).count() == 0:
                return CampaignAPI.objects.create(campaign=campaign_obj)

    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("check_and_create_campaign_api: %s %s", str(e), str(exc_tb.tb_lineno), extra={
            'AppName': 'Campaign'})


def check_and_create_campaign_users(trigger_setting_obj):
    try:
        campaign_batch_data = json.loads(trigger_setting_obj.campaign.batch.sample_data)

        for data in campaign_batch_data:
            if not CampaignVoiceUser.objects.filter(mobile_number=data[0], voice_campaign=trigger_setting_obj):
                CampaignVoiceUser.objects.create(mobile_number=data[0], voice_campaign=trigger_setting_obj)
        
    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("check_and_create_campaign_users: %s %s", str(e), str(exc_tb.tb_lineno), extra={'AppName': 'Campaign'})


def get_whatsapp_audience_log_data(data, bot_obj, is_export):
    try:
        campaign_id = data["campaign_ids"][0]
        status_filter = data["status_filter"]
        quick_reply_filter = data["quick_reply"]
        searched_type = data["searched_type"]
        searched_value = data['searched_value']
        test_message_filter = data.get('test_message_filter', list())
        all_campaign_objs = []
        selected_campaigns = []
        
        status_list = [CAMPAIGN_COMPLETED, CAMPAIGN_FAILED,
                       CAMPAIGN_ONGOING, CAMPAIGN_PARTIALLY_COMPLETED, CAMPAIGN_DRAFT]

        campaign_obj = Campaign.objects.filter(
            pk=campaign_id, bot=bot_obj)

        audience_log_objs = CampaignAudienceLog.objects.filter(
            campaign=campaign_id)
        
        if len(test_message_filter) == 1:
            if test_message_filter[0].strip().lower() == 'test_message':
                audience_log_objs = audience_log_objs.filter(is_test=True)
            elif test_message_filter[0].strip().lower() == 'campaign_message':
                audience_log_objs = audience_log_objs.filter(is_test=False)

        if len(status_filter) == 1:
            if status_filter[0].strip().lower() == 'failed':
                audience_log_objs = audience_log_objs.filter(is_failed=True)
            elif status_filter[0].strip().lower() == 'success':
                audience_log_objs = audience_log_objs.filter(is_failed=False)
        if quick_reply_filter:
            quick_reply_query = reduce(operator.or_, (Q(quick_replies__name=qr) for qr in quick_reply_filter))
            audience_log_objs = audience_log_objs.filter(quick_reply_query)

        if searched_type == 'phone_number':
            audience_log_objs = audience_log_objs.filter(audience__audience_id=searched_value)
        elif searched_type == 'recipient_id':
            audience_log_objs = audience_log_objs.filter(recepient_id=searched_value)
        
        selected_template_names = list(campaign_obj.filter(campaign_template__isnull=False).values_list(
            "campaign_template__template_name", flat=True).distinct())
        
        if data["is_first_render"]:
            camapign_status_query = reduce(
                operator.or_, (Q(status=status) for status in status_list))
            all_campaign_objs = Campaign.objects.filter(channel__value="whatsapp", bot=bot_obj,
                                                        is_deleted=False)
            all_campaign_objs = list(all_campaign_objs.filter(
                camapign_status_query).values_list("pk", "name").distinct())

        return_data = {'selected_template_names': selected_template_names,
                       'selected_campaigns': selected_campaigns,
                       'all_campaign_objs': all_campaign_objs,
                       'searched_value': searched_value,
                       'searched_type': searched_type,
                       'quick_reply_filter': ', '.join(quick_reply_filter)}
        if is_export:
            return_data['campaign_names'] = ', '.join(set(campaign_obj.values_list('name', flat=True)))

        return audience_log_objs, return_data
    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("get_whatsapp_audience_log_data: %s %s", str(e), str(exc_tb.tb_lineno), extra={'AppName': 'Campaign'})

    return CampaignAudienceLog.objects.none(), {}


# def get_whatsapp_audience_logs(data, bot_obj, is_export):
#     try:
#         campaign_id = data["campaign_ids"]
#         # filter_date_type = data["filter_date_type"]
#         status_filter = data["status_filter"]
#         quick_reply_filter = data["quick_reply"]
#         template_names = data['templates']
#         # start_date = bot_obj.created_datetime.date()
#         searched_type = data["searched_type"]
#         searched_value = data['searched_value']
#         # end_date = datetime.now().date()
#         is_template_select = data['is_template_select']
#         all_campaign_objs = []
#         selected_campaigns = []
        
#         # if filter_date_type == "1":
#         #     start_date = (datetime.now() - timedelta(days=1)).date()
#         # elif filter_date_type == "2":
#         #     start_date = (datetime.now() - timedelta(days=7)).date()
#         # elif filter_date_type == "3":
#         #     start_date = (datetime.now() - timedelta(days=30)).date()
#         # elif filter_date_type == "5":
#         #     start_date = datetime.strptime(data["start_date"], "%Y-%m-%d").date()
#         #     end_date = datetime.strptime(data["end_date"], "%Y-%m-%d").date()
        
#         status_list = [CAMPAIGN_COMPLETED, CAMPAIGN_FAILED,
#                        CAMPAIGN_ONGOING, CAMPAIGN_PARTIALLY_COMPLETED]
#         if not is_template_select:
#             camapign_status_query = reduce(operator.or_, (Q(status=status) for status in status_list))

#             # camapign_ids_query = reduce(operator.or_, (Q(pk=camp_id) for camp_id in campaign_ids))
#             campaign_obj = Campaign.objects.filter(
#                 pk=campaign_id[0], bot=bot_obj, channel__value="whatsapp", is_deleted=False)
#             campaign_obj = campaign_obj.filter(camapign_status_query)
#             # if template_names:
#             #     template_name_query = reduce(operator.or_, (Q(campaign_template__template_name=template) for template in template_names))
#             #     campaign_objs = campaign_objs.filter(template_name_query)

#             audience_log_objs = CampaignAudienceLog.objects.filter(
#                 campaign=campaign_obj)

#             # audience_log_objs = audience_log_objs.filter(created_date__gte=start_date, created_date__lte=end_date)

#             # if campaign_ids and not audience_log_objs:
#             #     return CampaignAudienceLog.objects.none(), {"unauthorised_campaign": "unauthorised_campaign"}
#         else:
#             campaign_obj = Campaign.objects.filter(
#                 bot=bot_obj, channel__value="whatsapp", campaign_template__template_name__in=template_names, status__in=status_list, is_deleted=False)
#             audience_log_objs = CampaignAudienceLog.objects.filter(
#                 campaign=campaign_obj)
#             # audience_log_objs = audience_log_objs.filter(created_date__gte=start_date, created_date__lte=end_date)
#             selected_campaigns = list(
#                 campaign_obj.values_list("pk", flat=True))

#         if len(status_filter) == 1:
#             if status_filter[0].strip().lower() == 'failed':
#                 audience_log_objs = audience_log_objs.filter(is_failed=True)
#             elif status_filter[0].strip().lower() == 'success':
#                 audience_log_objs = audience_log_objs.filter(is_failed=False)
#         if quick_reply_filter:
#             quick_reply_query = reduce(operator.or_, (Q(quick_replies__name=qr) for qr in quick_reply_filter))
#             audience_log_objs = audience_log_objs.filter(quick_reply_query)

#         if searched_type == 'phone_number':
#             audience_log_objs = audience_log_objs.filter(audience__audience_id=searched_value)
#         elif searched_type == 'recipient_id':
#             audience_log_objs = audience_log_objs.filter(recepient_id=searched_value)
        
#         selected_template_names = list(campaign_obj.filter(campaign_template__isnull=False).values_list(
#             "campaign_template__template_name", flat=True).distinct())
        
#         if data["is_first_render"]:
#             all_campaign_objs = list(Campaign.objects.filter(channel__value="whatsapp", bot=bot_obj,
#                                      status__in=status_list, is_deleted=False).values_list("pk", "name").distinct())

#         return_data = {'selected_template_names': selected_template_names,
#                        'selected_campaigns': selected_campaigns,
#                        'all_campaign_objs': all_campaign_objs,
#                        'searched_value': searched_value,
#                        'searched_type': searched_type,
#                        'quick_reply_filter': ', '.join(quick_reply_filter)}
#         if is_export:
#             return_data['campaign_names'] = ', '.join(set(audience_log_objs.values_list('campaign__name', flat=True)))
#         return audience_log_objs, return_data
#     except Exception as e:
#         exc_type, exc_obj, exc_tb = sys.exc_info()
#         logger.error("get_whatsapp_audience_logs: %s %s", str(e), str(exc_tb.tb_lineno), extra={'AppName': 'Campaign'})

#     return CampaignAudienceLog.objects.none(), {}


def get_whatsapp_audience_user_data(audience_log_obj, quick_replies_obj):
    try:
        _, status, _, error_title = add_status_for_bsp(
            audience_log_obj.campaign, audience_log_obj)
        time_zone = settings.TIME_ZONE
        quick_reply = ', '.join(list(quick_replies_obj.filter(
            audience_log=audience_log_obj).values_list('name', flat=True)))

        request = json.loads(audience_log_obj.request)
        
        template = request.get('template')
        if template:
            template_name = template.get('name')
            template_code = template.get('language', {}).get('code', '')
            template = template_name + ' (' + template_code + ')'
        else:
            campaign_template = audience_log_obj.campaign.campaign_template
            if campaign_template is not None:
                template = campaign_template.template_name + ' ' + \
                    campaign_template.language.title
            else:
                template = '-'

        return {
            'phone_number': audience_log_obj.audience.audience_id,
            'recipient_id': audience_log_obj.recepient_id if audience_log_obj.recepient_id else '-',
            'status': status,
            'campaign': audience_log_obj.campaign.name,
            'template': template,
            'failure_reason': error_title,
            'quick_reply': quick_reply if quick_reply else '-',
            'sent_time': audience_log_obj.sent_datetime.astimezone(pytz.timezone(time_zone)).strftime("%d-%m-%Y %H:%M:%S") if audience_log_obj.is_sent else '-',
            'delivered_time': audience_log_obj.delivered_datetime.astimezone(pytz.timezone(time_zone)).strftime("%d-%m-%Y %H:%M:%S") if audience_log_obj.is_delivered else '-',
            'read_time': audience_log_obj.read_datetime.astimezone(pytz.timezone(time_zone)).strftime("%d-%m-%Y %H:%M:%S") if audience_log_obj.is_read else '-',
            'chat_history': "Click Here",
            'is_test': audience_log_obj.is_test,
        }
    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("get_whatsapp_audience_user_data: %s %s", str(
            e), str(exc_tb.tb_lineno), extra={'AppName': 'Campaign'})

    return {}


def get_whatsapp_users_stats(campaign_audience_objs):
    user_stats_list = {}
    try:
        user_stats_list = campaign_audience_objs.values("campaign__bot").annotate(
            total_sent=Count('pk', filter=Q(is_sent=True, is_test=False)),
            total_delivered=Count('pk', filter=Q(is_delivered=True, is_test=False)),
            total_read=Count('pk', filter=Q(is_read=True, is_test=False)),
            total_failed=Count('pk', filter=Q(is_failed=True, is_test=False)),
            total_replied=Count('pk', filter=Q(is_replied=True, is_test=False)),
            total_test_sent=Count('pk', filter=Q(is_sent=True, is_test=True)),
            total_test_failed=Count('pk', filter=Q(is_failed=True, is_test=True)),
            total_tested=Count('pk', filter=Q(is_test=True)),
            total_processed=Count('pk', filter=Q(is_processed=True)))

    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("get_whatsapp_users_stats: %s %s", str(e), str(
            exc_tb.tb_lineno), extra={'AppName': 'Campaign'})

    return user_stats_list


def get_whatsapp_audience_reports(audience_log_objs, data):
    filename = []
    try:
        user_stats = get_whatsapp_users_stats(audience_log_objs)
        campaign_details = get_campaign_details_data(audience_log_objs[0].campaign)
        whatsapp_bsp_name = campaign_details["whatsapp_bsp_name"]
        if whatsapp_bsp_name == "-":
            whatsapp_bsp_name = "Ameyo"

        total_audience = audience_log_objs.count()
        total_sheet = math.ceil(total_audience / CAMPAIGN_REPORT_MAX_ROWS)
        sheet_count = 1
        start_audience_index = 0
        while (sheet_count <= total_sheet):
            test_wb = openpyxl.Workbook(write_only=True)
            sheet1 = test_wb.create_sheet("Whatsapp Overall report")
            sheet2 = test_wb.create_sheet("Whatsapp Detailed report")
            
            add_whatsapp_overall_details(
                sheet1, user_stats, campaign_details, data)

            audience_log_objs = audience_log_objs.values('id', 'audience__audience_id', 'recepient_id', 'is_sent', 'sent_datetime', 'is_delivered',
                                                         'delivered_datetime', 'is_read', 'read_datetime', 'is_replied', 'replied_datetime', 'is_failed', 'is_test', 'request', 'response')

            add_default_whatsapp_audience_history_columns(sheet2)
            for audience_log_obj in audience_log_objs[start_audience_index:CAMPAIGN_REPORT_MAX_ROWS * sheet_count].iterator():
                add_sheet_whatsapp_audience_history_report(audience_log_obj, sheet2, whatsapp_bsp_name)

            filepath = 'wpm_reports/WhatsappCampaignHistory' + \
                '-' + campaign_details["campaign_name"] + '-' + str(datetime.now()) + '.xlsx'
            test_wb.save(settings.MEDIA_ROOT + filepath)
            filename.append(filepath)
            start_audience_index = CAMPAIGN_REPORT_MAX_ROWS * sheet_count
            sheet_count += 1

    except Exception as exc:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("get_whatsapp_audience_reports: %s at %s", str(exc), str(
            exc_tb.tb_lineno), extra={'AppName': 'Campaign'})

    return filename


def get_voice_campaign_user_data(voice_campaign_user_obj):
    try:
        disposition_code_data = voice_campaign_user_obj.get_dispostion_code_data()

        return {
            "voice_campaign_user_id": voice_campaign_user_obj.pk,
            "phone_number": voice_campaign_user_obj.mobile_number,
            "status": voice_campaign_user_obj.get_call_status(),
            "campaign_name": voice_campaign_user_obj.voice_campaign.campaign.name,
            "date_created": voice_campaign_user_obj.get_created_date_time(),
            "type": "Outbound",
            "duration": voice_campaign_user_obj.get_call_duration(),
            "disposition_code": disposition_code_data[0],
            "disposition_intent": disposition_code_data[1],
            "disposition_tree": disposition_code_data[2],
            "disposition_parent_tree": disposition_code_data[3],
            "setup_time": "N/A",
            "ring_time": "N/A",
            "hang_up": "N/A",
            "recording": voice_campaign_user_obj.get_call_recording(),
            "transcript": voice_campaign_user_obj.get_voice_user_id()
        }
        
    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("get_voice_campaign_user_data: %s %s", str(e), str(exc_tb.tb_lineno), extra={'AppName': 'Campaign'})

    return {}


def update_voice_campaign_user_call_status(voice_campaign_obj):
    try:
        voice_campaign_user_objs = CampaignVoiceUser.objects.filter(voice_campaign=voice_campaign_obj, status=None)
        for voice_campaign_user_obj in voice_campaign_user_objs:
            voice_campaign_user_obj.status = "failed"
            voice_campaign_user_obj.save()
            
    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("update_voice_campaign_user_call_status: %s %s", str(e), str(exc_tb.tb_lineno), extra={'AppName': 'Campaign'})


def clients_data_build_dict(seq, key):
    try:
        return dict((int(validation_obj.removing_phone_non_digit_element(d[key])) if key == 'phone_number' else d[key], dict(d)) for d in seq)
    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("clients_data_build_dict: %s %s", str(e), str(exc_tb.tb_lineno), extra={'AppName': 'Campaign'})
        return {}


def get_audience_unique_id(unique_id):
    if not isinstance(unique_id, str):
        unique_id = str(unique_id)
    unique_id = unique_id.strip()
    unique_id = remo_html_from_string(unique_id)
    unique_id = unique_id
    return unique_id
    # elif isinstance(unique_id, dict): for future scalability
    #     json_unique_id = {}
    #     for id in unique_id:
    #         key = str(id)
    #         key = remo_html_from_string(key.strip())
    #         value = str(unique_id[id])
    #         value = remo_html_from_string(value.strip())
    #         json_unique_id[key] = value
    #     json_unique_id = json.dumps(json_unique_id)


def check_pipe_delimiter_in_csv(csv_header):
    try:
        header = csv_header[0]
        if '|' in header:
            return True
    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("In check_pipe_delimiter_in_csv: %s %s", str(e), str(
            exc_tb.tb_lineno), extra={'AppName': 'Campaign'})

    return False


def get_pipe_separated_row_data_list(row_data):
    try:
        psv_row_data = row_data[0].split('|')
        return psv_row_data
    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("In get_pipe_separated_row_data_list: %s %s", str(e), str(
            exc_tb.tb_lineno), extra={'AppName': 'Campaign'})

    return []


def get_csv_sheet_details(csv_sheet, delimiter):
    max_row, max_column = 0, 0
    try:
        csv_reader = csv.reader(csv_sheet, delimiter=delimiter)
        csv_header = next(csv_reader)
        is_pipe = check_pipe_delimiter_in_csv(csv_header)
        if is_pipe:
            csv_header = get_pipe_separated_row_data_list(csv_header)
        max_column = len(csv_header)
        max_row += 1 + sum([1 for row in csv_reader if row])

    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("In get_csv_sheet: %s %s", str(e), str(
            exc_tb.tb_lineno), extra={'AppName': 'Campaign'})

    return max_row, max_column, is_pipe


def csv_sample_data_wrapper_function(file_path, Profile, campaign_channel, original_file_name, event_obj, bot_obj, status_message, metadata):
    try:
        try:
            func_timeout.func_timeout(
                CAMPAIGN_MAXIMUM_SHEET_UPLOAD_TIME_CSV_LIMIT, get_sample_data_from_csv_batch_file, args=[file_path, Profile, campaign_channel, original_file_name, event_obj, bot_obj, status_message, metadata])
        except func_timeout.FunctionTimedOut:
            os.remove(file_path)
            status = 400
            row_errors = [
                "The process took longer than expected. Please try again later."]
            total_batch_count = metadata['max_row'] - 1
            event_obj.event_info = json.dumps({"row_errors": row_errors,
                                               "status": status,
                                               "original_file_name": original_file_name,
                                               "header_metadata": [],
                                               "sample_data": [],
                                               "total_batch_count": total_batch_count,
                                               "total_opted_in": 0,
                                               "status_message": status_message
                                               })
            event_obj.is_completed = True
            event_obj.event_progress = 0
            event_obj.save()

    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("In csv_sample_data_wrapper_function: %s at %s", str(e), str(exc_tb.tb_lineno), extra={
            'AppName': 'Campaign'})


def check_running_upload_events(user, bot, event_type, max_event_count):
    try:
        if event_type == "upload_batch":
            current_time = timezone.now().astimezone(pytz.timezone(settings.TIME_ZONE))
            create_datetime_diff = current_time - \
                timedelta(
                    minutes=CAMPAIGN_MAXIMUM_SHEET_UPLOAD_TIME_CSV_LIMIT // 60 + 20)

            CampaignEventProgress.objects.filter(
                user=user, bot=bot, event_type=event_type, event_datetime__lte=create_datetime_diff, is_completed=False).update(is_completed=True)
            running_event_objs = CampaignEventProgress.objects.filter(
                user=user, bot=bot, event_type=event_type, is_completed=False)
            if running_event_objs.count() > max_event_count - 1:
                return True

    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("In check_running_upload_events: %s at %s", str(e), str(exc_tb.tb_lineno), extra={
            'AppName': 'Campaign'})

    return False


def create_campaign_event_progress_tracker_obj(user, bot, event_type):
    event_obj = None
    try:
        event_obj = CampaignEventProgress.objects.create(
            user=user,
            bot=bot,
            event_type=event_type,
        )
    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("In create_campaign_event_progress_tracker_obj: %s at %s", str(e), str(exc_tb.tb_lineno), extra={
            'AppName': 'Campaign'})
    
    return event_obj


def get_sample_data_from_csv_batch_file(file_path, Profile, campaign_channel, original_file_name, event_obj, bot_obj, status_message, metadata):
    try:
        csv_sheet = open(file_path, 'r+')
        delimiter = metadata['delimiter']
        csv_reader = csv.reader(csv_sheet, delimiter=delimiter)
        row_errors, header_metadata, sample_data, status = [], [], [], None
        max_row, max_column = metadata['max_row'], metadata['max_column']
        auto_delete_checked = metadata['auto_delete_checked']
        error_rows = metadata['error_rows']
        error_found, total_opted_in, deleted_rows = False, 0, 0
        csv_header = next(csv_reader)

        first_col_name = csv_header[0]
        header_data = []
        for col in range(0, max_column):
            header_name = csv_header[col]
            if not header_name:
                header_name = ''
            header_name = remo_html_from_string(header_name)
            header_metadata.append({
                'col_no': col,
                'col_name': header_name.replace(',', '')
            })
            header_data.append(header_name)

        if 'phone' not in first_col_name.lower():
            status_message = 'First column should always be for the Phone.  Please change the column name ' + first_col_name + \
                ' to Phone or you can download and use our Template Batch file to upload the contacts in the prescribed text format only'
            status = 402
        elif '' in header_data:
            if (header_data.count('') > 1):
                status_message = "Two or more header names are empty or have no values filled in the columns of the uploaded file. Please enter the header names or remove the columns if not required and upload the file again."
            else:
                status_message = f"Empty header name detected in a column number {str(header_data.index('') + 1)} of the uploaded file. Please enter the header name or remove the column if not required and upload the file again."
            status = 402
        elif len(set(header_data)) != len(header_data):
            duplicate_column = list(
                set([x for x in header_data if header_data.count(x) > 1]))
            if (len(duplicate_column) > 1):
                status_message = f"We have detected the same header names {' and '.join(duplicate_column)} in the columns of the uploaded file. Please change the repeated header names with the unique names and upload the file again."
            else:
                status_message = f"We have detected the same header name {duplicate_column[0]} in the column of the uploaded file. Please change the repeated header name with a unique name and upload the file again."
            status = 402
        else:
            if auto_delete_checked:
                os.remove(file_path)
                metadata_for_deletion = {
                    "error_rows": error_rows, "csv_header": csv_header, "csv_reader": csv_reader, "max_row": max_row, "max_column": max_column, "delimiter": delimiter}

                changed_rows, sample_data, total_opted_in = delete_invalid_rows_of_csv_sheet(
                    file_path, event_obj, Profile, metadata_for_deletion)
                deleted_rows = len(error_rows)
                max_row = changed_rows
                error_rows = []

            else:
                error_file_name = original_file_name.split('.')[0] + ".csv"
                filename = generate_random_key(
                    10) + "_" + error_file_name.replace(" ", "")
                default_file_path = "secured_files/CampaignApp/campaign_batch/" + filename
                row_processed = 1

                with open(default_file_path, 'w') as csv_error_sheet:
                    csv_writer = csv.writer(csv_error_sheet, delimiter=",")
                    csv_writer.writerow(csv_header)

                    for row in csv_reader:
                        if not row:
                            row_processed += 1
                            continue

                        progress = int((row_processed - 1) /
                                       (max_row - 1) * 100)
                        if row_processed % 500 == 0:
                            event_obj.event_progress = progress
                            event_obj.save(update_fields=["event_progress"])

                        mobile_number = row[0]
                        mobile_number, error_found = identify_invalid_phone_numbers(
                            mobile_number, campaign_channel, row_processed)
                        if error_found:
                            error_rows.append(row_processed)
                            row_errors.append(error_found)
                            row.append(error_found)
                            csv_writer.writerow(row)

                        profile_obj = Profile.objects.filter(
                            user_id=mobile_number)
                        if profile_obj and profile_obj[0].campaign_optin:
                            total_opted_in += 1

                        # To get sample batch data to show on modal
                        if not error_found and row_processed < 4:
                            row_data_value = [mobile_number]
                            for col_index in range(1, max_column):
                                value = row[col_index] if col_index < len(
                                    row) else '-'
                                row_data_value.append(clean_input(
                                    value) if value and value != '-' else '-')
                            sample_data.append(row_data_value)

                        row_processed += 1
            
            csv_sheet.close()

        if max_row < 2:
            status = 403
        total_batch_count = max_row - 1
        response = {}
        if row_errors:
            file_access_management_obj = CampaignFileAccessManagement.objects.create(
                file_path=default_file_path, is_public=False, original_file_name=original_file_name, bot=bot_obj)
            response["file_path"] = "/campaign/download-file/" + \
                str(file_access_management_obj.key) + "/"
        else:
            file_access_management_obj = CampaignFileAccessManagement.objects.create(
                file_path=file_path, is_public=False, original_file_name=original_file_name, bot=bot_obj)
            response["file_path"] = "/campaign/download-file/" + \
                str(file_access_management_obj.key) + "/"

        if row_errors or status in [402, 403]:
            if not status:
                status = 400
            row_errors = row_errors[:CAMPAIGN_MAXIMUM_SCROLLER_LIST_LIMIT]
            event_obj.event_info = json.dumps({"row_errors": row_errors,
                                               "status": status,
                                               "original_file_name": original_file_name,
                                               "header_metadata": header_metadata,
                                               "sample_data": sample_data,
                                               "total_batch_count": total_batch_count,
                                               "total_opted_in": total_opted_in,
                                               "file_path": response["file_path"],
                                               "error_rows": error_rows,
                                               "status_message": status_message
                                               })
        else:
            status = 200
            message = "success"
            if campaign_channel == "RCS":
                rcs_enabled_users = 0
                rcs_enabled_users, status, status_message = add_rcs_enabled_column(
                    file_path, bot_obj, RCSDetails)
                if status == 200:
                    total_opted_in = rcs_enabled_users

            event_obj.event_info = json.dumps({"row_errors": row_errors[:CAMPAIGN_MAXIMUM_SCROLLER_LIST_LIMIT],
                                               "status": status,
                                               "original_file_name": original_file_name,
                                               "header_metadata": header_metadata,
                                               "sample_data": sample_data,
                                               "total_batch_count": total_batch_count,
                                               "total_opted_in": total_opted_in,
                                               "file_path": response["file_path"],
                                               "error_rows": error_rows,
                                               "status_message": status_message,
                                               "deleted_rows": deleted_rows,
                                               "message": message
                                               })
        event_obj.is_completed = True
        event_obj.event_progress = 100
        event_obj.completed_datetime = datetime.now()
        event_obj.save()

    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("In get_sample_data_from_csv_batch_file: %s at %s", str(e), str(exc_tb.tb_lineno), extra={
            'AppName': 'Campaign'})

        event_obj.is_completed = True
        event_obj.completed_datetime = datetime.now()
        event_obj.save()


def delete_invalid_rows_of_csv_sheet(file_path, event_obj, Profile, metadata_for_deletion):
    try:
        sample_data = []
        total_opted_in = 0
        csv_header = metadata_for_deletion["csv_header"]
        csv_reader = metadata_for_deletion["csv_reader"]
        error_rows = metadata_for_deletion["error_rows"]
        max_row = metadata_for_deletion["max_row"]
        max_column = metadata_for_deletion["max_column"]
        delimiter = metadata_for_deletion["delimiter"]

        with open(file_path, 'w') as csv_updated_sheet:
            csv_writer = csv.writer(csv_updated_sheet, delimiter=delimiter)
            csv_header2 = []
            col = 1
            for header in csv_header:
                if header:
                    csv_header2.append(header)
                else:
                    csv_header2.append('Value ' + str(col))
                    col += 1

            csv_writer.writerow(csv_header2)
            row_processed, changed_rows = 1, 1

            for row in csv_reader:
                if not row:
                    row_processed += 1
                    continue
                progress = int((row_processed - 1) /
                               (max_row - 1) * 100)
                if row_processed % 500 == 0:
                    event_obj.event_progress = progress
                    event_obj.save(update_fields=["event_progress"])
                if row_processed not in error_rows:
                    changed_rows += 1
                    csv_writer.writerow(row)
                    mobile_number = str(row[0])
                    profile_obj = Profile.objects.filter(
                        user_id=mobile_number)
                    if profile_obj and profile_obj[0].campaign_optin:
                        total_opted_in += 1

                    if changed_rows < 5:
                        mobile_number = clean_input(mobile_number)
                        mobile_number = int(validation_obj.removing_phone_non_digit_element(mobile_number))
                        row_data_value = [mobile_number]
                        for col_index in range(1, max_column):
                            value = row[col_index] if col_index < len(
                                row) else '-'
                            row_data_value.append(clean_input(
                                value) if value and value != '-' else '-')
                        sample_data.append(row_data_value)
                row_processed += 1

    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("In delete_invalid_rows_of_csv_sheet: %s at %s", str(e), str(exc_tb.tb_lineno), extra={
            'AppName': 'Campaign'})

    return changed_rows, sample_data, total_opted_in


# checking whether audience log have data present for the campaign(s) been exported
def get_audience_log_exists_for_export_of_campaign(single_campaign, campaign_obj, custom_encrypt_obj, response):
    try:
        audience_log_exists = CampaignAudienceLog.objects.filter(campaign=campaign_obj).exists()
        if not audience_log_exists:
            response["status"] = 400
            if single_campaign:
                response["message"] = "There is no data available to export for the selected campaign."
            else:
                response["message"] = "There is no data available to export for the selected date range."
            response = json.dumps(response)
            encrypted_response = custom_encrypt_obj.encrypt(
                response)
            response = {"Response": encrypted_response}
            return True, response
    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("In get_audience_log_exists_for_export_of_campaign: %s at %s", str(e), str(exc_tb.tb_lineno), extra={
            'AppName': 'Campaign'})
    return False, response


# This function updates the statutes to schedule completed of the scheduled campaign if those are completed
def update_schedule_status():
    try:
        cronjob_detector_id = "update_schedule_status_cronjob"
        expiration_time = CAMPAIGN_ANALYTICS_SCHEDULAR_EXPIRY_TIME

        campaign_cronjob_tracker_obj = get_campaign_cronjob_tracker_obj(cronjob_detector_id)
        if campaign_cronjob_tracker_obj:
            if campaign_cronjob_tracker_obj.is_object_expired(expiration_time):
                delete_campaign_cronjob_tracker_obj(cronjob_detector_id)
                create_campaign_cronjob_tracker_obj(cronjob_detector_id)
            else:
                logger.info("Campaign App update schedule service is already running!",
                            extra={'AppName': 'CampaignApp'})
                return
        else:
            create_campaign_cronjob_tracker_obj(cronjob_detector_id)
        scheduled_campaigns_objs = Campaign.objects.filter(status=CAMPAIGN_SCHEDULED, is_deleted=False).iterator()
        for scheduled_campaigns_obj in scheduled_campaigns_objs:
            campaign_schedule_obj = CampaignScheduleObject.objects.filter(campaign=scheduled_campaigns_obj, is_sent=False).exists()
            if not campaign_schedule_obj:
                scheduled_campaigns_obj.status = CAMPAIGN_SCHEDULE_COMPLETED
                scheduled_campaigns_obj.save(update_fields=['status'])

    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("In update_schedule_status: %s at %s", str(e), str(exc_tb.tb_lineno), extra={
            'AppName': 'Campaign'})
    delete_campaign_cronjob_tracker_obj(cronjob_detector_id)


# Updates CAMPAIGN_IN_PROGRESS Campaign Analytics
def update_inprogress_campaigns_analytics():
    try:
        cronjob_detector_id = INPROGRESS_CAMPAIGN_ANALYTICS_SCHEDULAR_ID_CONSTANT
        expiration_time = CAMPAIGN_ANALYTICS_SCHEDULAR_EXPIRY_TIME

        campaign_cronjob_tracker_obj = get_campaign_cronjob_tracker_obj(
            cronjob_detector_id)
        if campaign_cronjob_tracker_obj:
            if campaign_cronjob_tracker_obj.is_object_expired(expiration_time):
                delete_campaign_cronjob_tracker_obj(cronjob_detector_id)
                create_campaign_cronjob_tracker_obj(cronjob_detector_id)
            else:
                logger.info("Campaign App update analytics service is already running!",
                            extra={'AppName': 'CampaignApp'})
                return
        else:
            create_campaign_cronjob_tracker_obj(cronjob_detector_id)

        bot_obj = Bot.objects.filter(is_deleted=False)
        current_time = timezone.now()
        # finding all the campaigns in the past two days that are in progress for whatsapp
        start_time = current_time - timedelta(days=2)

        campaign_objs = Campaign.objects.filter(
            Q(is_deleted=False),
            Q(bot__in=bot_obj), Q(channel__value='whatsapp'),
            Q(status__in=[CAMPAIGN_IN_PROGRESS,
              CAMPAIGN_PROCESSED, CAMPAIGN_DRAFT, CAMPAIGN_ONGOING]),
            Q(start_datetime__date__gte=start_time)
        )

        for campaign_obj in campaign_objs.iterator():
            try:
                analytics_obj = CampaignAnalytics.objects.get(
                    campaign=campaign_obj)
            except Exception:
                continue
            is_draft = campaign_obj.status == CAMPAIGN_DRAFT

            analytics_metadata = update_campaign_analytics_data(
                campaign_obj, analytics_obj)
            if analytics_metadata and not is_draft:
                update_campaign_status(campaign_obj, analytics_metadata)

    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("Error: update_inprogress_campaigns_analytics: %s for %s", str(
            e), str(exc_tb.tb_lineno), extra={'AppName': 'Campaign'})

    delete_campaign_cronjob_tracker_obj(cronjob_detector_id)


# Updates all Campaigns Analytics except In_progress campaigns
def update_campaigns_analytics():
    try:
        bot_obj = Bot.objects.filter(is_deleted=False)
        current_time = timezone.now()
        # finding all the campaigns in the past two days that are in progress for whatsapp
        start_time = current_time - timedelta(days=2)

        campaign_objs = Campaign.objects.filter(
            Q(is_deleted=False),
            Q(bot__in=bot_obj),
            Q(channel__value='whatsapp'),
            ~Q(status__in=[CAMPAIGN_IN_PROGRESS,
               CAMPAIGN_PROCESSED, CAMPAIGN_DRAFT, CAMPAIGN_SCHEDULED, CAMPAIGN_SCHEDULE_COMPLETED]),
            Q(start_datetime__date__gte=start_time)
        )

        for campaign_obj in campaign_objs.iterator():
            try:
                analytics_obj = CampaignAnalytics.objects.get(
                    campaign=campaign_obj)
            except Exception:
                continue

            analytics_metadata = update_campaign_analytics_data(
                campaign_obj, analytics_obj)
            if analytics_metadata:
                update_campaign_status(campaign_obj, analytics_metadata)

    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("Error: update_campaigns_analytics: %s for %s", str(
            e), str(exc_tb.tb_lineno), extra={'AppName': 'Campaign'})


def update_campaign_analytics_data(campaign_obj, analytics_obj):
    try:

        audience_log_objs = CampaignAudienceLog.objects.filter(
            campaign=campaign_obj)
        user_stats_list = get_whatsapp_users_stats(audience_log_objs)
        if user_stats_list.exists():
            user_stats_list = user_stats_list[0]
            message_sent = user_stats_list['total_sent']
            unsuccessful = user_stats_list['total_failed']
            processed = user_stats_list['total_processed']
            total_tested = user_stats_list['total_tested']
            analytics_obj.message_sent = message_sent
            analytics_obj.message_delivered = user_stats_list['total_delivered']
            analytics_obj.message_read = user_stats_list['total_read']
            analytics_obj.message_replied = user_stats_list['total_replied']
            analytics_obj.message_unsuccessful = unsuccessful
            analytics_obj.test_message_sent = user_stats_list['total_test_sent']
            analytics_obj.test_message_unsuccessful = user_stats_list['total_test_failed']
            analytics_obj.total_tested = total_tested
            analytics_obj.message_processed = processed
            analytics_obj.save(update_fields=[
                'test_message_sent', 'test_message_unsuccessful', 'total_tested', 'message_sent', 'message_delivered', 'message_read', 'message_replied', 'message_unsuccessful', 'message_processed'])

            analytics_metadata = {"message_sent": message_sent, "total_audience": analytics_obj.total_audience,
                                  "unsuccessful": unsuccessful, "processed": processed, 'total_tested': total_tested}

            return analytics_metadata

    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("In update_campaign_analytics_data: %s at %s", str(e), str(exc_tb.tb_lineno), extra={
            'AppName': 'Campaign'})
        return {}


def update_campaign_status(campaign_obj, analytics_metadata):
    try:
        if campaign_obj.status == CAMPAIGN_ONGOING:
            return
        unsuccessful = analytics_metadata["unsuccessful"]
        message_sent = analytics_metadata["message_sent"]
        processed_message = analytics_metadata["processed"]
        total_audience = analytics_metadata["total_audience"]
        total_tested = analytics_metadata["total_tested"]

        if processed_message - total_tested > 0 and (processed_message - total_tested) % total_audience == 0:
            if campaign_obj.status == CAMPAIGN_IN_PROGRESS:
                campaign_obj.processed_datetime = datetime.now()
            campaign_obj.status = CAMPAIGN_PROCESSED

            if unsuccessful == 0 and message_sent > 0 and (message_sent % total_audience == 0):
                campaign_obj.status = CAMPAIGN_COMPLETED
            elif message_sent == 0 and unsuccessful > 0 and (unsuccessful % total_audience == 0):
                campaign_obj.status = CAMPAIGN_FAILED
            elif (message_sent + unsuccessful) % total_audience == 0 and message_sent > 0 and unsuccessful > 0:
                campaign_obj.status = CAMPAIGN_PARTIALLY_COMPLETED

        else:
            campaign_obj.status = CAMPAIGN_IN_PROGRESS
        
        campaign_obj.save(update_fields=['status', 'processed_datetime'])

    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("In update_campaign_status: %s at %s", str(e), str(exc_tb.tb_lineno), extra={
            'AppName': 'Campaign'})


def validate_aws_sqs_credentials(bot_wsp_obj, campaign_wsp_config_meta):
    try:
        aws_key_id, aws_secret_access_key, aws_sqs = bot_wsp_obj.aws_key_id, bot_wsp_obj.aws_secret_access_key, bot_wsp_obj.aws_sqs
        sqs_client = boto3.client(
            'sqs', 'ap-south-1', aws_access_key_id=aws_key_id, aws_secret_access_key=aws_secret_access_key)

        campaign_wsp_config_meta["aws_sqs"] = aws_sqs
        campaign_wsp_config_meta["sqs_client"] = sqs_client

        sqs_domain = f'{settings.EASYCHAT_HOST_URL}/campaign/lambda/push-message/'
        sqs_packet = {
            "url": sqs_domain,
            "campaign_id": "",
            "campaign_api_id": "",
            "audience_id": "",
            "parameter": {},
            "event_name": "INITIAL_PUSH_TRIGGER_EVENT"
        }
        sqs_response = send_message_into_campaign_sqs(
            json.dumps(sqs_packet), sqs_client, aws_sqs)

        return sqs_response

    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("In validate_aws_sqs_credentials: %s at %s", str(e), str(exc_tb.tb_lineno), extra={
            'AppName': 'Campaign'})

        return None


def get_value_from_cache(key, dynamic_prefix):
    try:
        value = cache.get(key + str(dynamic_prefix))
        return value
    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("get_value_from_cache: %s %s", str(e), str(exc_tb.tb_lineno), extra={'AppName': 'Campaign'})
        return None


def set_value_to_cache(key, dynamic_prefix, value):
    try:
        cache.set(key + str(dynamic_prefix), value, settings.CACHE_TIME)
    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("set_value_to_cache: %s %s", str(e), str(exc_tb.tb_lineno), extra={'AppName': 'Campaign'})


def get_whatsapp_api_access_token(bot):
    API_KEY = ""
    try:
        whatsapp_credential_config_obj = WhatsappCredentialsConfig.objects.filter(bot=bot).first()
        if whatsapp_credential_config_obj:
            domain = whatsapp_credential_config_obj.host_url
            username = whatsapp_credential_config_obj.username
            password = whatsapp_credential_config_obj.password
            logger.info("=== Inside AMEYO_GET_API_KEY API ===",
                        extra={'AppName': 'Campaign'})
            url = domain + "/v1/users/login"
            userAndPass = base64.b64encode(
                (username + ":" + password).encode()).decode("ascii")
            headers = {
                'Content-Type': 'application/json',
                'Authorization': 'Basic ' + userAndPass
            }
            logger.info("AMEYO_GET_API_KEY Headers: %s", str(
                headers), extra={'AppName': 'Campaign'})
            request = requests.request("POST", url, headers=headers,
                                       timeout=25, verify=True)
            content = json.loads(request.text)
            logger.info("AMEYO_GET_API_KEY Response: %s", str(
                content), extra={'AppName': 'Campaign'})
            if str(request.status_code) == "200":
                if "users" in content and content["users"] != []:
                    API_KEY = content["users"][0]["token"]
            logger.info("AMEYO_GET_API_KEY API Key generated: %s",
                        str(API_KEY), extra={'AppName': 'Campaign'})
    except requests.Timeout as RT:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("AMEYO_GET_API_KEY Timeout error: %s at %s", str(
            RT), str(exc_tb.tb_lineno), extra={'AppName': 'Campaign'})
    except Exception as E:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("AMEYO_GET_API_KEY Failed: %s at %s", str(
            E), str(exc_tb.tb_lineno), extra={'AppName': 'Campaign'})
    return API_KEY


def get_queue_attributes(attribute_names, bot_wsp_obj):
    response = ""
    try:
        aws_key_id, aws_secret_access_key, queue_url = bot_wsp_obj.aws_key_id, bot_wsp_obj.aws_secret_access_key, bot_wsp_obj.aws_sqs
        sqs_client = boto3.client(
            'sqs', 'ap-south-1', aws_access_key_id=aws_key_id, aws_secret_access_key=aws_secret_access_key)

        response = sqs_client.get_queue_attributes(
            QueueUrl=queue_url, AttributeNames=attribute_names)
    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("Error in get_queue_attributes: %s %s", str(
            e), str(exc_tb.tb_lineno), extra={'AppName': 'Campaign'})

    return response


def get_messages_in_queue(bot_wsp_obj):
    try:
        bot_obj = bot_wsp_obj.bot
        current_time = timezone.now()
        # Finding all the campaigns in the past two days that are in progress for whatsapp
        start_time = current_time - timedelta(days=2)
        campaign_objs = Campaign.objects.filter(Q(is_deleted=False), Q(bot=bot_obj), Q(
            channel__value='whatsapp'), Q(start_datetime__date__gte=start_time), status=CAMPAIGN_IN_PROGRESS)
        total_audience, total_message_processed = 0, 0
        for campaign_obj in campaign_objs.iterator():
            try:
                analytics_obj = CampaignAnalytics.objects.get(
                    campaign=campaign_obj)
                total_batch_audience = analytics_obj.total_audience
                message_processed = max(
                    (analytics_obj.message_processed - analytics_obj.total_tested), 0)
                total_message_processed += message_processed
                total_audience += total_batch_audience
            except Exception:
                continue
        total_message_in_queue = total_audience - total_message_processed
        return total_message_in_queue
    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("Error in get_messages_in_queue: %s %s", str(
            e), str(exc_tb.tb_lineno), extra={'AppName': 'Campaign'})
        return 0


def is_livechat_connected(phone_number, bot_obj):
    try:
        user_objs = Profile.objects.filter(
            user_id=phone_number, bot=bot_obj, livechat_connected=True)
        if user_objs.exists():
            return True

    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("Error in is_livechat_connected: %s %s", str(
            e), str(exc_tb.tb_lineno), extra={'AppName': 'Campaign'})
    return False


def livechat_active_session_failure_response(audience_log_obj):
    audience_log_obj.is_failed = True
    failure_response = {"errors": [{"code": 500, "title": "Failed to send the message because the user was already in an active LiveChat session.",
                                    "details": "The message was not sent to this user because there was already an active LiveChat session during the same time period. You can try to resend the to this user again after some time."}]}
    audience_log_obj.request = json.dumps({})
    audience_log_obj.response = json.dumps(failure_response)
    audience_log_obj.is_processed = True
    audience_log_obj.save(
        update_fields=["request", "response", "is_failed", "is_processed"])
    return failure_response


def get_campaign_wsp_config_meta(campaign_obj, bot_wsp_id, response, is_test=False):
    try:
        campaign_wsp_config_meta = {}
        campaign_api = CampaignAPI.objects.filter(
            campaign=campaign_obj)

        bot_wsp_obj = CampaignBotWSPConfig.objects.get(pk=bot_wsp_id)

        if not campaign_api or not campaign_api[0].is_api_completed:
            response[
                'message'] = 'Send Campaign failed because API Integration is pending.'
        else:
            campaign_api = campaign_api[0]
            campaign_api.campaign_bot_wsp_config = bot_wsp_obj
            campaign_api.save()

            campaign_wsp_config_meta = {"code": bot_wsp_obj.code, "namespace": bot_wsp_obj.namespace,
                                        "enable_queuing_system": False if is_test else bot_wsp_obj.enable_queuing_system, "bot_wsp_id": bot_wsp_id}
            response['status'] = 201
            
    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("Error in get_campaign_wsp_config_meta: %s %s", str(
            e), str(exc_tb.tb_lineno), extra={'AppName': 'Campaign'})
    return response, campaign_wsp_config_meta, bot_wsp_obj


def create_cloned_campaign(campaign_obj, campaign_name, new_batch_obj=None):
    try:
        channel_obj = campaign_obj.channel
        batch_obj = new_batch_obj if new_batch_obj else campaign_obj.batch
        cloned_campaign_obj = Campaign.objects.create(
            name=campaign_name,
            channel=campaign_obj.channel,
            bot=campaign_obj.bot,
            status=CAMPAIGN_DRAFT,
            last_saved_state=CAMPAIGN_BASIC_INFO_STATE,
            campaign_template=campaign_obj.campaign_template,
            batch=batch_obj,
            parent_campaign_id=campaign_obj.pk,
        )

        existing_template_variable_obj = CampaignTemplateVariable.objects.filter(
            campaign=campaign_obj, template=campaign_obj.campaign_template, batch=campaign_obj.batch).first()
        if existing_template_variable_obj:
            existing_template_variable_obj.pk = None
            existing_template_variable_obj.campaign = cloned_campaign_obj
            existing_template_variable_obj.batch = batch_obj
            existing_template_variable_obj.save()

        cloned_campaign_api_obj = check_and_create_campaign_api(
            channel_obj, cloned_campaign_obj, CampaignAPI, CampaignVoiceBotAPI)
        existing_campaign_wsp = CampaignAPI.objects.filter(
            campaign=campaign_obj).first().campaign_bot_wsp_config
        cloned_campaign_api_obj.campaign_bot_wsp_config = existing_campaign_wsp
        cloned_campaign_api_obj.save(
            update_fields=['campaign_bot_wsp_config'])
        return cloned_campaign_obj
    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("Error in create_cloned_campaign: %s %s", str(
            e), str(exc_tb.tb_lineno), extra={'AppName': 'Campaign'})


def verify_selected_batch_for_scheduling(campaign_obj, new_batch_pk, response):
    try:
        new_campaign_batch = None
        new_campaign_batch = CampaignBatch.objects.filter(pk=new_batch_pk).first()
    
        old_batch_obj = campaign_obj.batch
        old_batch_header_meta = json.loads(old_batch_obj.batch_header_meta)
        new_batch_header_meta = json.loads(
            new_campaign_batch.batch_header_meta)

        old_batch_header_meta_len = len(old_batch_header_meta)
        new_batch_header_meta_len = len(new_batch_header_meta)

        if old_batch_header_meta_len != new_batch_header_meta_len:
            response["status"] = 400
        else:
            response["status"] = 200
            for old_header_idx in range(old_batch_header_meta_len):
                for new_header_idx in range(old_header_idx, new_batch_header_meta_len):
                    if old_batch_header_meta[old_header_idx]['col_no'] == new_batch_header_meta[new_header_idx]['col_no']:
                        if old_batch_header_meta[old_header_idx]['col_name'] != new_batch_header_meta[new_header_idx]['col_name']:
                            response["status"] = 400
                        break
                if response["status"] == 400:
                    break
    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        logger.error("Error in verify_selected_batch_for_scheduling: %s %s", str(
            e), str(exc_tb.tb_lineno), extra={'AppName': 'Campaign'})
        response["status"] = 500
    return new_campaign_batch


def make_response_packet(response, custom_encrypt_obj):
    response = json.dumps(response)
    encrypted_response = custom_encrypt_obj.encrypt(response)
    response = {"Response": encrypted_response}
    return Response(data=response)
