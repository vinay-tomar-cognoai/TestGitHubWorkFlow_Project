from openpyxl import Workbook, load_workbook
import os


def update_message_history_excel_headings():
    dirs = os.listdir("files/message-history/")
    
    for dr in dirs:
        if "bot-" in dr:
            for excel_name in ["/MessageHistoryLastOneDay.xlsx", "/MessageHistoryLastSevenDays.xlsx", "/MessageHistoryLastThirtyDays.xlsx"]:
                try:
                    wb = load_workbook("files/message-history/" + dr + excel_name)
                    sheet = wb.worksheets[0]
                    sheet.column_dimensions['O'].width = 15
                    sheet.cell(row=1, column=15).value = "Type-in Query"
                    wb.save("files/message-history/" + dr + excel_name)
                except Exception:
                    pass

update_message_history_excel_headings()
